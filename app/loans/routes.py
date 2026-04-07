"""Loan management routes"""
from flask import render_template, redirect, url_for, flash, request, current_app, jsonify, make_response
from flask_login import login_required, current_user
import io
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import os
from app import db
from app.loans import loans_bp
from app.models import Loan, LoanPayment, Customer, ActivityLog, SystemSettings, User, LoanScheduleOverride, Branch
from app.loans.forms import LoanForm, LoanPaymentForm, EditPaymentForm, LoanApprovalForm, StaffApprovalForm, ManagerApprovalForm, InitiateLoanForm, AdminApprovalForm, LoanDeactivationForm
from app.utils.decorators import permission_required, admin_required
from app.utils.helpers import generate_loan_number, generate_customer_id, get_current_branch_id, should_filter_by_branch, generate_receipt_number


def _calculate_loan_totals_for_principal(
    principal_amount,
    interest_rate,
    loan_type,
    interest_type,
    duration_months=None,
    duration_weeks=None,
    duration_days=None,
):
    """Calculate installment amount and total payable from a principal base amount."""
    from decimal import Decimal, ROUND_HALF_UP, ROUND_UP

    principal = Decimal(str(principal_amount or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    rate = Decimal(str(interest_rate or 0))
    months = int(duration_months or 0)
    weeks = int(duration_weeks or 0)
    days = int(duration_days or 0)

    if principal <= 0:
        return Decimal('0.00'), Decimal('0.00')

    if loan_type == 'type1_9weeks':
        weeks = weeks or 9
        interest = rate * Decimal('2')
        total_payable = (principal * (Decimal('100') + interest) / Decimal('100')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        emi = (total_payable / Decimal(str(weeks))).quantize(Decimal('1'), rounding=ROUND_UP)
    elif loan_type == '54_daily':
        days = days or 54
        full_interest = rate * Decimal('2')
        total_payable = (principal * (Decimal('100') + full_interest) / Decimal('100')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        emi = (total_payable / Decimal(str(days))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    elif loan_type == 'type4_micro':
        months = months or 1
        weeks = weeks or (months * 4)
        full_interest = rate * Decimal(str(months))
        emi = (principal * ((full_interest + Decimal('100')) / Decimal('100')) / Decimal(str(weeks))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        total_payable = (emi * Decimal(str(weeks))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    elif loan_type == 'type4_daily':
        months = months or 1
        days = days or (months * 26)
        full_interest = rate * Decimal(str(months))
        emi = (principal * ((full_interest + Decimal('100')) / Decimal('100')) / Decimal(str(days))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        total_payable = (emi * Decimal(str(days))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    elif loan_type == 'special_loan':
        total_interest = (principal * rate / Decimal('100')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        total_payable = (principal + total_interest).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        emi = total_payable
    else:
        months = months or 1
        monthly_rate = rate / (Decimal('12') * Decimal('100'))
        if interest_type == 'reducing_balance' and monthly_rate > 0:
            mr_float = float(monthly_rate)
            power_calc = ((1 + mr_float) ** months) / (((1 + mr_float) ** months) - 1)
            emi = (principal * monthly_rate * Decimal(str(power_calc))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            total_interest = principal * rate * Decimal(str(months)) / (Decimal('12') * Decimal('100'))
            emi = ((principal + total_interest) / Decimal(str(months))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        total_payable = (emi * Decimal(str(months))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    return emi, total_payable


def _refresh_loan_financial_state(loan):
    """Refresh derived loan financial fields after principal/disbursement edits."""
    from decimal import Decimal, ROUND_HALF_UP

    # 1) Recalculate current outstanding using model rules
    loan.update_outstanding_amount()

    # 2) Recalculate advance balance from schedule allocation (keeps loan card
    # and schedule "Advance out" in sync, even with skipped installments).
    schedule = loan.generate_payment_schedule()
    loan.advance_balance = loan.calculate_available_advance_balance(schedule=schedule)

    # 3) Rebuild historical `balance_after` in payment history from new total payable
    running_outstanding = Decimal(str(loan.total_payable or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    payments = loan.payments.order_by(LoanPayment.payment_date.asc(), LoanPayment.id.asc()).all()
    for payment in payments:
        pay_amount = Decimal(str(payment.payment_amount or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        running_outstanding = (running_outstanding - pay_amount).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        if running_outstanding < Decimal('0.00'):
            running_outstanding = Decimal('0.00')
        payment.balance_after = float(running_outstanding)

    # 4) Keep status aligned with outstanding
    current_outstanding = loan.calculate_current_outstanding()
    if current_outstanding <= Decimal('0.02'):
        loan.status = 'completed'
        if not loan.closing_date:
            last_payment = loan.payments.order_by(LoanPayment.payment_date.desc(), LoanPayment.id.desc()).first()
            loan.closing_date = last_payment.payment_date if last_payment else datetime.utcnow().date()
    elif loan.status == 'completed':
        loan.status = 'active'
        loan.closing_date = None


def _get_installment_advance_breakdown(loan):
    """Return installment/advance deduction amounts for payment collection UI."""
    from decimal import Decimal, ROUND_HALF_UP

    installment_amount = Decimal(str(loan.installment_amount or 0)).quantize(
        Decimal('0.01'),
        rounding=ROUND_HALF_UP,
    )
    advance_balance = loan.calculate_available_advance_balance()

    if installment_amount < Decimal('0.00'):
        installment_amount = Decimal('0.00')
    if advance_balance < Decimal('0.00'):
        advance_balance = Decimal('0.00')

    advance_to_apply = min(advance_balance, installment_amount)
    deducted_installment_amount = (installment_amount - advance_to_apply).quantize(
        Decimal('0.01'),
        rounding=ROUND_HALF_UP,
    )

    return {
        'installment_amount': installment_amount,
        'advance_balance': advance_balance,
        'advance_to_apply': advance_to_apply,
        'deducted_installment_amount': deducted_installment_amount,
    }


def _resolve_payment_amount_with_optional_advance(
    posted_amount,
    use_advance_credit,
    installment_amount,
    advance_to_apply,
):
    """Resolve effective cash collection amount for installment payment screen."""
    from decimal import Decimal, ROUND_HALF_UP

    amount = Decimal(str(posted_amount or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    installment = Decimal(str(installment_amount or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    advance = Decimal(str(advance_to_apply or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    if amount < Decimal('0.00'):
        amount = Decimal('0.00')
    if installment < Decimal('0.00'):
        installment = Decimal('0.00')
    if advance < Decimal('0.00'):
        advance = Decimal('0.00')

    advance = min(advance, installment)
    deducted_installment = (installment - advance).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    if not use_advance_credit or advance <= Decimal('0.00'):
        return amount, False

    # Auto-convert only when collector entered exact installment amount.
    if installment > Decimal('0.00') and abs(amount - installment) <= Decimal('0.05'):
        return deducted_installment, True

    # Keep collector-entered custom amount untouched.
    return amount, False


def _get_active_user_for_customer(customer):
    """Return active system user mapped to a customer by NIC, if any."""
    if not customer or not customer.nic_number:
        return None

    return User.query.filter(
        User.nic_number == customer.nic_number,
        User.is_active == True
    ).first()


def _resolve_staff_loan_customer_selection(selected_customer_id):
    """Resolve customer for staff loan selection.

    Supports existing customer ids and synthetic negative ids representing users.
    For synthetic user ids, creates a member profile on demand if missing.
    """
    selected_raw = str(selected_customer_id or '').strip()
    if not selected_raw:
        return None, 'Please select a staff member for Staff Loan.'

    try:
        selected_id = int(selected_raw)
    except (TypeError, ValueError):
        return None, 'Invalid staff member selection.'

    current_branch_id = get_current_branch_id() if should_filter_by_branch() else None

    # Existing member selected
    if selected_id > 0:
        customer = Customer.query.get(selected_id)
        if not customer:
            return None, 'Selected member not found.'

        if current_branch_id and customer.branch_id != current_branch_id:
            return None, 'Selected member is not in the current branch.'

        if not _get_active_user_for_customer(customer):
            return None, 'Staff Loan is only available for members linked to an active user in Settings > Users (NIC must match).'

        # Normalize legacy auto-created staff-linked records to proxy type.
        if customer.is_staff_member_profile and 'staff_user_proxy' not in customer.customer_types:
            customer.customer_types = ['staff_user_proxy']
            db.session.flush()

        return customer, None

    # Synthetic user selection (negative user id)
    user_id = abs(selected_id)
    user = User.query.filter(
        User.id == user_id,
        User.is_active == True
    ).first()
    if not user:
        return None, 'Selected staff user is not active or not found.'

    if current_branch_id and user.branch_id != current_branch_id:
        return None, 'Selected staff user is not in the current branch.'

    # Reuse existing member profile by NIC when present
    customer = Customer.query.filter_by(nic_number=user.nic_number).first()
    if customer:
        if current_branch_id and customer.branch_id != current_branch_id:
            return None, 'Linked member record is not in the current branch.'
        if customer.status != 'active':
            return None, 'Linked member record is not active.'

        # Normalize legacy auto-created records to internal proxy type.
        if customer.is_staff_member_profile and 'staff_user_proxy' not in customer.customer_types:
            customer.customer_types = ['staff_user_proxy']
            db.session.flush()

        return customer, None

    # Auto-create a minimal member profile for this user
    branch_id = user.branch_id or current_branch_id
    if not branch_id:
        fallback_branch = Branch.query.filter_by(is_active=True).order_by(Branch.id.asc()).first()
        branch_id = fallback_branch.id if fallback_branch else None

    if not branch_id:
        return None, 'No active branch found to create member profile for this user.'

    generated_customer_id = generate_customer_id('customer', branch_id)
    customer = Customer(
        customer_id=generated_customer_id,
        branch_id=branch_id,
        full_name=user.full_name,
        nic_number=user.nic_number,
        phone_primary=user.phone or 'N/A',
        email=user.email,
        address_line1='Auto-created from Settings User',
        address_line2='',
        city='N/A',
        district='Colombo',
        status='active',
        kyc_verified=False,
        created_by=current_user.id,
        notes=f'Auto-created for Staff Loan from settings user: {user.username}'
    )
    customer.customer_types = ['staff_user_proxy']
    db.session.add(customer)
    db.session.flush()
    return customer, None


def _resolve_staff_loan_guarantor_ids(raw_guarantor_ids, borrower_customer_id):
    """Resolve and validate guarantor ids for staff loans."""
    raw_tokens = [token.strip() for token in str(raw_guarantor_ids or '').split(',') if token.strip()]
    if len(raw_tokens) != 2:
        return None, 'Staff Loan requires exactly 2 staff guarantors.'

    resolved_ids = []
    for token in raw_tokens:
        guarantor_customer, error = _resolve_staff_loan_customer_selection(token)
        if error:
            return None, f'Invalid staff guarantor selection: {error}'

        if guarantor_customer.id == borrower_customer_id:
            return None, 'Borrower cannot be selected as a guarantor.'

        if guarantor_customer.id in resolved_ids:
            return None, 'Duplicate guarantor selected. Please choose two different staff guarantors.'

        resolved_ids.append(guarantor_customer.id)

    return resolved_ids, None

@loans_bp.route('/')
@login_required
@permission_required('manage_loans')
def list_loans():
    """List all loans"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)
    search = request.args.get('search', '')
    status = request.args.get('status', '')
    loan_type = request.args.get('loan_type', '')
    referred_by = request.args.get('referred_by', type=int)
    
    query = Loan.query
    
    # Filter by current branch if needed
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id:
            query = query.filter_by(branch_id=current_branch_id)
    
    if search:
        query = query.join(Customer).filter(
            db.or_(
                Loan.loan_number.ilike(f'%{search}%'),
                Customer.full_name.ilike(f'%{search}%'),
                Customer.customer_id.ilike(f'%{search}%')
            )
        )
    
    if status:
        query = query.filter_by(status=status)
    
    if loan_type:
        query = query.filter_by(loan_type=loan_type)
    
    if referred_by:
        query = query.filter_by(referred_by=referred_by)
    
    loans = query.order_by(Loan.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Get users for referrer filter
    user_query = User.query.filter_by(is_active=True)
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id:
            user_query = user_query.filter_by(branch_id=current_branch_id)
    users = user_query.order_by(User.full_name).all()
    
    return render_template('loans/list.html',
                         title='Loans',
                         loans=loans,
                         search=search,
                         status=status,
                         loan_type=loan_type,
                         referred_by=referred_by,
                         users=users)

@loans_bp.route('/add', methods=['GET', 'POST'])
@login_required
@permission_required('manage_loans')
def add_loan():
    """Add new loan"""
    form = LoanForm()
    
    # Get users for referred_by dropdown
    user_query = User.query.filter_by(is_active=True)
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id:
            user_query = user_query.filter_by(branch_id=current_branch_id)
    
    users = user_query.order_by(User.full_name).all()
    form.referred_by.choices = [(0, 'Select User (Optional)')] + [(u.id, f'{u.full_name} ({u.username})') for u in users]
    
    # Get admin and accountant users as final approver candidates
    final_approver_query = User.query.filter(
        User.is_active == True,
        User.role.in_(['admin', 'accountant'])
    )
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id:
            final_approver_query = final_approver_query.filter(
                db.or_(User.role == 'admin', User.branch_id == current_branch_id)
            )
    final_approvers = final_approver_query.order_by(User.role.desc(), User.full_name).all()
    
    # Pre-fill interest rate from settings on GET request
    if request.method == 'GET':
        settings = SystemSettings.get_settings()
        form.interest_rate.data = settings.default_loan_interest_rate
        form.interest_type.data = settings.interest_calculation_method
        form.duration_months.data = settings.default_loan_duration
        form.installment_frequency.data = 'monthly'  # Default for monthly loans
    
    if form.validate_on_submit():
        # Custom validation based on loan type
        if form.loan_type.data == 'type1_9weeks':
            # For Type 1 loans, validate weeks instead of months
            if not form.duration_weeks.data:
                flash('Duration (Weeks) is required for Type 1 - 9 Week Loan!', 'error')
                return render_template('loans/add.html', title='Add Loan', form=form, final_approvers=final_approvers)
        elif form.loan_type.data == '54_daily':
            # For 54 Daily loans, validate days instead of months
            if not form.duration_days.data:
                flash('Duration (Days) is required for 54 Daily Loan!', 'error')
                return render_template('loans/add.html', title='Add Loan', form=form, final_approvers=final_approvers)
        elif form.loan_type.data == 'type4_micro':
            # For Type 4 Micro loans, validate months (will convert to weeks internally)
            if not form.duration_months.data:
                flash('Duration (Months) is required for Type 4 - Micro Loan!', 'error')
                return render_template('loans/add.html', title='Add Loan', form=form, final_approvers=final_approvers)
        elif form.loan_type.data == 'type4_daily':
            # For Type 4 Daily loans, validate months (will convert to days internally)
            if not form.duration_months.data:
                flash('Duration (Months) is required for Type 4 - Daily Loan!', 'error')
                return render_template('loans/add.html', title='Add Loan', form=form, final_approvers=final_approvers)
        elif form.loan_type.data == 'special_loan':
            # For Special Loans, validate start and end dates
            if not form.start_date.data:
                flash('Start Date is required for Special Loan!', 'error')
                return render_template('loans/add.html', title='Add Loan', form=form, final_approvers=final_approvers)
            if not form.end_date.data:
                flash('End Date is required for Special Loan!', 'error')
                return render_template('loans/add.html', title='Add Loan', form=form, final_approvers=final_approvers)
            if form.end_date.data <= form.start_date.data:
                flash('End Date must be after Start Date!', 'error')
                return render_template('loans/add.html', title='Add Loan', form=form, final_approvers=final_approvers)
        elif form.loan_type.data in ['monthly_loan', 'staff_loan']:
            # For Monthly/Staff loans, validate months and interest type
            if not form.duration_months.data:
                flash('Duration (Months) is required for Monthly/Staff Loan!', 'error')
                return render_template('loans/add.html', title='Add Loan', form=form, final_approvers=final_approvers)
            if not form.interest_type.data:
                flash('Interest Type is required for Monthly/Staff Loan!', 'error')
                return render_template('loans/add.html', title='Add Loan', form=form, final_approvers=final_approvers)
        else:
            # For other loan types, validate months, interest type, and installment frequency
            if not form.duration_months.data:
                flash('Duration (Months) is required!', 'error')
                return render_template('loans/add.html', title='Add Loan', form=form, final_approvers=final_approvers)
            if not form.interest_type.data:
                flash('Interest Type is required!', 'error')
                return render_template('loans/add.html', title='Add Loan', form=form, final_approvers=final_approvers)
            if not form.installment_frequency.data:
                flash('Installment Frequency is required!', 'error')
                return render_template('loans/add.html', title='Add Loan', form=form, final_approvers=final_approvers)
        
        # Validate customer selection
        if form.customer_id.data == 0:
            flash('Please select a customer!', 'error')
            return render_template('loans/add.html', title='Add Loan', form=form, final_approvers=final_approvers)
        
        # Resolve selected member
        if form.loan_type.data == 'staff_loan':
            customer, staff_error = _resolve_staff_loan_customer_selection(form.customer_id.data)
            if staff_error:
                flash(staff_error, 'error')
                return render_template('loans/add.html', title='Add Loan', form=form, final_approvers=final_approvers)
        else:
            customer = Customer.query.get(form.customer_id.data)
            if not customer:
                flash('Customer not found!', 'error')
                return render_template('loans/add.html', title='Add Loan', form=form, final_approvers=final_approvers)
        
        if not customer.branch_id:
            flash('Customer does not have a valid branch assigned!', 'error')
            return render_template('loans/add.html', title='Add Loan', form=form, final_approvers=final_approvers)

        raw_guarantor_ids = (request.form.get('guarantor_ids') or '').strip()
        if form.loan_type.data == 'staff_loan':
            resolved_guarantor_ids, guarantor_error = _resolve_staff_loan_guarantor_ids(
                raw_guarantor_ids,
                borrower_customer_id=customer.id
            )
            if guarantor_error:
                flash(guarantor_error, 'error')
                return render_template('loans/add.html', title='Add Loan', form=form, final_approvers=final_approvers)
            guarantor_ids_value = ','.join(str(gid) for gid in resolved_guarantor_ids)
        else:
            guarantor_ids_value = raw_guarantor_ids
        
        # Generate loan number with new format: YY/B##/TYPE/#####
        loan_number = generate_loan_number(loan_type=form.loan_type.data, branch_id=customer.branch_id)
        
        # Calculate EMI using Decimal arithmetic
        from decimal import Decimal, ROUND_HALF_UP, ROUND_DOWN, ROUND_UP
        
        loan_amount = Decimal(str(form.loan_amount.data))
        interest_rate = Decimal(str(form.interest_rate.data))
        
        # Determine duration and calculation method based on loan type
        duration_weeks = None
        duration_days = None
        duration_months = form.duration_months.data
        
        # Check if this is a Type 1 - 9 week loan
        if form.loan_type.data == 'type1_9weeks':
            duration_weeks = form.duration_weeks.data or 9
            duration_months = 0  # Not used for weekly loans
            # Type 1 calculation: Interest = Interest rate * 2
            # Total Payable = Loan Amount * (100 + Interest) / 100  ← computed first to avoid rounding drift
            # EMI = CEIL(Total Payable / weeks)  ← last installment absorbs remainder
            interest = interest_rate * Decimal('2')
            total_payable = (loan_amount * (Decimal('100') + interest) / Decimal('100')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            emi = (total_payable / Decimal(str(duration_weeks))).quantize(Decimal('1'), rounding=ROUND_UP)
        elif form.loan_type.data == '54_daily':
            duration_days = form.duration_days.data or 54
            duration_months = 0  # Not used for daily loans
            # 54 Daily Loan: Full Interest = Interest Rate * 2
            # Total Payable = Loan Amount * (100 + Full Interest) / 100
            # Daily Installment = Total Payable / Days
            full_interest = interest_rate * Decimal('2')
            total_payable = (loan_amount * (Decimal('100') + full_interest) / Decimal('100')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            emi = (total_payable / Decimal(str(duration_days))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        elif form.loan_type.data == 'type4_micro':
            # Type 4 Micro Loan: Uses months as input, converts to weeks
            # Full Interest = Interest Rate * Months
            # Weeks = Months * 4
            # Installment = LA * ((Full Interest + 100) / 100) / Weeks
            months = form.duration_months.data
            duration_weeks = months * 4
            full_interest = interest_rate * Decimal(str(months))
            emi = (loan_amount * ((full_interest + Decimal('100')) / Decimal('100'))) / Decimal(str(duration_weeks))
            emi = emi.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            total_payable = (emi * Decimal(str(duration_weeks))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        elif form.loan_type.data == 'type4_daily':
            # Type 4 Daily Loan: Uses months as input, converts to days (1 month = 26 days)
            # Full Interest = Interest Rate * Months
            # Days = Months * 26
            # Installment = LA * ((Full Interest + 100) / 100) / Days
            months = form.duration_months.data
            duration_days = months * 26
            full_interest = interest_rate * Decimal(str(months))
            emi = (loan_amount * ((full_interest + Decimal('100')) / Decimal('100'))) / Decimal(str(duration_days))
            emi = emi.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            total_payable = (emi * Decimal(str(duration_days))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        elif form.loan_type.data == 'special_loan':
            # Special Loan: No installments, full payment at end date
            # Total Payable = Loan Amount + (Loan Amount * Interest Rate / 100)
            duration_months = 0
            total_interest = (loan_amount * interest_rate / Decimal('100')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            total_payable = (loan_amount + total_interest).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            emi = total_payable  # Single payment at end date
        elif form.loan_type.data in ['monthly_loan', 'staff_loan']:
            # Monthly/Staff Loan: Standard monthly calculation
            monthly_rate = interest_rate / (Decimal('12') * Decimal('100'))
            n = duration_months
            
            if form.interest_type.data == 'reducing_balance' and monthly_rate > 0:
                # EMI = [P x R x (1+R)^N]/[(1+R)^N-1]
                mr_float = float(monthly_rate)
                power_calc = ((1 + mr_float) ** n) / (((1 + mr_float) ** n) - 1)
                emi = loan_amount * monthly_rate * Decimal(str(power_calc))
            else:
                # Flat rate calculation
                total_interest = loan_amount * interest_rate * Decimal(str(n)) / (Decimal('12') * Decimal('100'))
                emi = (loan_amount + total_interest) / Decimal(str(n))
            
            emi = emi.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            total_payable = (emi * Decimal(str(n))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            # Standard monthly calculation
            monthly_rate = interest_rate / (Decimal('12') * Decimal('100'))
            n = duration_months
            
            if form.interest_type.data == 'reducing_balance' and monthly_rate > 0:
                # Convert to float for power calculation, then back to Decimal
                mr_float = float(monthly_rate)
                power_calc = ((1 + mr_float) ** n) / (((1 + mr_float) ** n) - 1)
                emi = loan_amount * monthly_rate * Decimal(str(power_calc))
            else:
                total_interest = loan_amount * interest_rate * Decimal(str(n)) / (Decimal('12') * Decimal('100'))
                emi = (loan_amount + total_interest) / Decimal(str(n))
            
            emi = emi.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            total_payable = (emi * Decimal(str(n))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        # Handle document upload
        document_filename = None
        if form.document.data and hasattr(form.document.data, 'filename') and form.document.data.filename:
            file = form.document.data
            filename = secure_filename(file.filename)
            # Create unique filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            document_filename = f"{timestamp}_{filename}"
            
            # Create upload directory if it doesn't exist
            upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'loans')
            os.makedirs(upload_folder, exist_ok=True)
            
            # Save the file
            file_path = os.path.join(upload_folder, document_filename)
            file.save(file_path)
            
            # Store relative path
            document_filename = f"loans/{document_filename}"
        
        # Get documentation fee from form (manually entered)
        documentation_fee = Decimal(str(form.documentation_fee.data or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        # Disbursed amount will be calculated during approval (loan amount minus documentation fee)
        # For pending loans, these remain None until approved
        actual_disbursed_amount = None
        
        loan = Loan(
            loan_number=loan_number,
            customer_id=customer.id,
            branch_id=customer.branch_id,
            loan_type=form.loan_type.data,
            loan_purpose=form.loan_purpose.data if form.loan_purpose.data else None,
            loan_amount=form.loan_amount.data,
            interest_rate=form.interest_rate.data,
            interest_type=form.interest_type.data,
            duration_months=duration_months,
            duration_weeks=duration_weeks,
            duration_days=duration_days,
            installment_amount=emi,
            installment_frequency='one_time' if form.loan_type.data == 'special_loan' else ('daily' if duration_days else ('weekly' if duration_weeks else form.installment_frequency.data)),
            disbursed_amount=actual_disbursed_amount,
            total_payable=total_payable,
            outstanding_amount=None,  # Will be set during approval
            documentation_fee=documentation_fee,
            application_date=form.start_date.data if form.loan_type.data == 'special_loan' and form.start_date.data else datetime.now().date(),
            maturity_date=form.end_date.data if form.loan_type.data == 'special_loan' else None,
            purpose=form.purpose.data,
            security_details=form.security_details.data,
            document_path=document_filename,
            drive_link=form.drive_link.data or None,
            guarantor_ids=guarantor_ids_value,
            final_approver_id=request.form.get('final_approver_id', type=int) or None,
            status='pending',  # All new loans start as pending and go through approval workflow
            created_by=current_user.id,
            referred_by=form.referred_by.data if form.referred_by.data != 0 else None,
            notes=form.notes.data
        )
        
        # Note: Disbursement details will be set during admin approval stage
        
        db.session.add(loan)
        
        # Log activity
        log = ActivityLog(
            user_id=current_user.id,
            action='create_loan',
            entity_type='loan',
            description=f'Created loan: {loan.loan_number}',
            ip_address=request.remote_addr
        )
        db.session.add(log)
        
        db.session.commit()
        
        flash(f'Loan {loan.loan_number} created successfully!', 'success')
        return redirect(url_for('loans.view_loan', id=loan.id))
    
    return render_template('loans/add.html', title='Add Loan', form=form, final_approvers=final_approvers)

@loans_bp.route('/edit-loan-select')
@login_required
@admin_required
def edit_loan_select():
    """Select loan to edit (Admin only)"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    status = request.args.get('status', '', type=str)
    loan_type = request.args.get('loan_type', '', type=str)
    
    query = Loan.query
    
    # Filter by current branch if needed
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id:
            query = query.filter_by(branch_id=current_branch_id)
    
    if search:
        query = query.join(Customer).filter(
            db.or_(
                Loan.loan_number.ilike(f'%{search}%'),
                Customer.full_name.ilike(f'%{search}%'),
                Customer.customer_id.ilike(f'%{search}%')
            )
        )
    
    if status:
        query = query.filter_by(status=status)
    
    if loan_type:
        query = query.filter_by(loan_type=loan_type)
    
    loans = query.order_by(Loan.created_at.desc()).paginate(
        page=page, per_page=25, error_out=False
    )
    
    return render_template('loans/edit_select.html',
                         title='Edit Loan - Select Loan',
                         loans=loans,
                         search=search,
                         status=status,
                         loan_type=loan_type)

@loans_bp.route('/<int:id>')
@login_required
def view_loan(id):
    """View loan details"""
    loan = Loan.query.get_or_404(id)
    
    # Check branch access
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id and loan.branch_id != current_branch_id:
            flash('Access denied: Loan not found in current branch.', 'danger')
            return redirect(url_for('loans.list_loans'))
    
    # Calculate current outstanding amount with accrued interest
    current_outstanding = loan.calculate_current_outstanding()
    accrued_interest = loan.calculate_accrued_interest()
    
    # Update the loan's stored outstanding amount to reflect current calculation
    loan.update_outstanding_amount()
    loan.advance_balance = loan.calculate_available_advance_balance()
    db.session.commit()
    
    # Get guarantors
    guarantors = []
    guarantor_loans = {}  # {customer_id: {'active': [...], 'history': [...]}}
    if loan.guarantor_ids:
        guarantor_id_list = [int(gid.strip()) for gid in loan.guarantor_ids.split(',') if gid.strip()]
        if guarantor_id_list:
            guarantors = Customer.query.filter(Customer.id.in_(guarantor_id_list)).all()
            for g in guarantors:
                g_loans = Loan.query.filter_by(customer_id=g.id).filter(Loan.id != loan.id).order_by(Loan.created_at.desc()).all()
                guarantor_loans[g.id] = {
                    'active': [l for l in g_loans if l.status in ['active', 'initiated', 'pending', 'pending_staff_approval', 'pending_manager_approval']],
                    'history': [l for l in g_loans if l.status in ['completed', 'rejected', 'deactivated', 'defaulted']],
                }

    # Get arrears details
    arrears_details = loan.get_arrears_details()
    
    # Get advance balance
    advance_balance = float(loan.advance_balance or 0)
    
    # Get payment history for this loan
    payments = loan.payments.order_by(LoanPayment.payment_date.desc()).all()
    
    # Generate payment schedule for display
    schedule = loan.generate_payment_schedule()
    
    return render_template('loans/view.html',
                         title=f'Loan: {loan.loan_number}',
                         loan=loan,
                         guarantors=guarantors,
                         guarantor_loans=guarantor_loans,
                         current_outstanding=current_outstanding,
                         accrued_interest=accrued_interest,
                         arrears_details=arrears_details,
                         advance_balance=advance_balance,
                         payments=payments,
                         schedule=schedule)

@loans_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_loan(id):
    """Edit loan (Admin only)"""
    loan = Loan.query.get_or_404(id)
    
    form = LoanForm()
    
    # Get users for referred_by dropdown
    user_query = User.query.filter_by(is_active=True)
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id:
            user_query = user_query.filter_by(branch_id=current_branch_id)
    
    users = user_query.order_by(User.full_name).all()
    form.referred_by.choices = [(0, 'Select User (Optional)')] + [(u.id, f'{u.full_name} ({u.username})') for u in users]
    
    if request.method == 'GET':
        # Pre-populate form with existing loan data
        customer = loan.customer
        form.customer_search.data = f"{customer.full_name} ({customer.customer_id})"
        form.customer_id.data = loan.customer_id
        form.referred_by.data = loan.referred_by if loan.referred_by else 0
        form.application_date.data = loan.application_date
        form.loan_type.data = loan.loan_type
        form.loan_purpose.data = loan.loan_purpose
        form.loan_amount.data = loan.loan_amount
        form.disbursed_amount.data = loan.disbursed_amount
        form.documentation_fee.data = loan.documentation_fee
        form.duration_weeks.data = loan.duration_weeks
        form.duration_days.data = loan.duration_days
        form.duration_months.data = loan.duration_months
        form.interest_rate.data = loan.interest_rate
        form.interest_type.data = loan.interest_type
        form.installment_frequency.data = loan.installment_frequency
        form.purpose.data = loan.purpose
        form.security_details.data = loan.security_details
        form.notes.data = loan.notes
        # Pre-populate special loan dates
        if loan.loan_type == 'special_loan':
            form.start_date.data = loan.application_date
            form.end_date.data = loan.maturity_date
    
    if form.validate_on_submit():
        # Custom validation based on loan type
        if form.loan_type.data == 'type1_9weeks':
            if not form.duration_weeks.data:
                flash('Duration (Weeks) is required for Type 1 - 9 Week Loan!', 'error')
                return render_template('loans/edit.html', title='Edit Loan', form=form, loan=loan)
        elif form.loan_type.data == '54_daily':
            if not form.duration_days.data:
                flash('Duration (Days) is required for 54 Daily Loan!', 'error')
                return render_template('loans/edit.html', title='Edit Loan', form=form, loan=loan)
        elif form.loan_type.data in ['type4_micro', 'type4_daily', 'monthly_loan', 'staff_loan']:
            if not form.duration_months.data:
                flash('Duration (Months) is required for this loan type!', 'error')
                return render_template('loans/edit.html', title='Edit Loan', form=form, loan=loan)
        elif form.loan_type.data == 'special_loan':
            if not form.start_date.data:
                flash('Start Date is required for Special Loan!', 'error')
                return render_template('loans/edit.html', title='Edit Loan', form=form, loan=loan)
            if not form.end_date.data:
                flash('End Date is required for Special Loan!', 'error')
                return render_template('loans/edit.html', title='Edit Loan', form=form, loan=loan)
            if form.end_date.data <= form.start_date.data:
                flash('End Date must be after Start Date!', 'error')
                return render_template('loans/edit.html', title='Edit Loan', form=form, loan=loan)
        
        # Validate customer selection
        if form.customer_id.data == 0:
            flash('Please select a customer!', 'error')
            return render_template('loans/edit.html', title='Edit Loan', form=form, loan=loan)
        
        # Resolve selected member
        if form.loan_type.data == 'staff_loan':
            customer, staff_error = _resolve_staff_loan_customer_selection(form.customer_id.data)
            if staff_error:
                flash(staff_error, 'error')
                return render_template('loans/edit.html', title='Edit Loan', form=form, loan=loan)
        else:
            customer = Customer.query.get(form.customer_id.data)
            if not customer:
                flash('Customer not found!', 'error')
                return render_template('loans/edit.html', title='Edit Loan', form=form, loan=loan)
        
        # Recalculate EMI and totals with updated values
        from decimal import Decimal, ROUND_HALF_UP
        
        loan_amount = Decimal(str(form.loan_amount.data))
        interest_rate = Decimal(str(form.interest_rate.data))

        edited_disbursed_amount = None
        if form.disbursed_amount.data is not None:
            edited_disbursed_amount = Decimal(str(form.disbursed_amount.data)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            if edited_disbursed_amount > loan_amount:
                flash('Disbursed Amount cannot be greater than Loan Amount.', 'error')
                return render_template('loans/edit.html', title='Edit Loan', form=form, loan=loan)
        
        duration_weeks = None
        duration_days = None
        duration_months = form.duration_months.data
        
        # Calculate based on loan type
        if form.loan_type.data == 'type1_9weeks':
            duration_weeks = form.duration_weeks.data or 9
            duration_months = 0
        elif form.loan_type.data == '54_daily':
            duration_days = form.duration_days.data or 54
            duration_months = 0
        elif form.loan_type.data == 'type4_micro':
            months = form.duration_months.data
            duration_weeks = months * 4
        elif form.loan_type.data == 'type4_daily':
            months = form.duration_months.data
            duration_days = months * 26
        elif form.loan_type.data == 'special_loan':
            duration_months = 0

        principal_for_calculation = edited_disbursed_amount if edited_disbursed_amount is not None else loan_amount
        emi, total_payable = _calculate_loan_totals_for_principal(
            principal_amount=principal_for_calculation,
            interest_rate=interest_rate,
            loan_type=form.loan_type.data,
            interest_type=form.interest_type.data,
            duration_months=duration_months,
            duration_weeks=duration_weeks,
            duration_days=duration_days,
        )

        # Always apply recalculated installment when disbursed amount changes so
        # Payment Schedule `Installment` and `Remaining` reflect the edited amount.
        effective_installment_amount = emi
        
        # Handle document upload
        if form.document.data and hasattr(form.document.data, 'filename') and form.document.data.filename:
            file = form.document.data
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            document_filename = f"{timestamp}_{filename}"
            
            upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'loans')
            os.makedirs(upload_folder, exist_ok=True)
            
            file_path = os.path.join(upload_folder, document_filename)
            file.save(file_path)
            
            loan.document_path = f"loans/{document_filename}"
        
        # Get documentation fee from form (manually entered)
        documentation_fee = Decimal(str(form.documentation_fee.data or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        # Update loan data
        loan.customer_id = form.customer_id.data
        loan.branch_id = customer.branch_id
        loan.loan_type = form.loan_type.data
        loan.loan_purpose = form.loan_purpose.data
        loan.loan_amount = form.loan_amount.data
        loan.interest_rate = form.interest_rate.data
        loan.interest_type = form.interest_type.data
        loan.duration_months = duration_months
        loan.duration_weeks = duration_weeks
        loan.duration_days = duration_days
        loan.installment_amount = effective_installment_amount
        loan.installment_frequency = 'one_time' if form.loan_type.data == 'special_loan' else ('daily' if duration_days else ('weekly' if duration_weeks else form.installment_frequency.data))
        loan.total_payable = total_payable
        loan.documentation_fee = documentation_fee
        if edited_disbursed_amount is not None:
            loan.disbursed_amount = edited_disbursed_amount
        elif loan.status in ['active', 'disbursed', 'completed'] and not loan.disbursed_amount:
            loan.disbursed_amount = loan.loan_amount
        loan.application_date = form.start_date.data if form.loan_type.data == 'special_loan' and form.start_date.data else form.application_date.data
        loan.maturity_date = form.end_date.data if form.loan_type.data == 'special_loan' else None
        loan.purpose = form.purpose.data
        loan.security_details = form.security_details.data
        loan.referred_by = form.referred_by.data if form.referred_by.data != 0 else None
        loan.notes = form.notes.data
        loan.updated_at = datetime.utcnow()
        
        # Recalculate all dependent financial fields for already disbursed loans
        if loan.status in ['active', 'disbursed', 'completed']:
            _refresh_loan_financial_state(loan)
        
        # Log activity
        log = ActivityLog(
            user_id=current_user.id,
            action='edit_loan',
            entity_type='loan',
            description=f'Edited loan: {loan.loan_number}',
            ip_address=request.remote_addr
        )
        db.session.add(log)
        
        db.session.commit()
        
        flash(f'Loan {loan.loan_number} updated successfully!', 'success')
        return redirect(url_for('loans.view_loan', id=loan.id))
    
    return render_template('loans/edit.html', title='Edit Loan', form=form, loan=loan)

@loans_bp.route('/<int:id>/approve', methods=['GET', 'POST'])
@login_required
@permission_required('approve_loans')
def approve_loan(id):
    """Approve loan"""
    loan = Loan.query.get_or_404(id)
    
    # Check branch access
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id and loan.branch_id != current_branch_id:
            flash('Access denied: Loan not found in current branch.', 'danger')
            return redirect(url_for('loans.list_loans'))
    
    if loan.status != 'pending':
        flash('Only pending loans can be approved!', 'warning')
        return redirect(url_for('loans.view_loan', id=id))
    
    # Check if this loan has a designated final approver
    if loan.final_approver_id and current_user.id != loan.final_approver_id:
        designated = User.query.get(loan.final_approver_id)
        name = designated.full_name if designated else 'Unknown'
        flash(f'This loan can only be finally approved by the designated approver: {name}.', 'danger')
        return redirect(url_for('loans.view_loan', id=id))
    
    form = LoanApprovalForm()
    
    if form.validate_on_submit():
        loan.approval_date = form.approval_date.data
        loan.approved_by = current_user.id
        loan.approval_notes = form.approval_notes.data
        
        if form.approval_status.data == 'approved':
            from decimal import Decimal
            loan.status = 'active'
            loan.approved_amount = form.approved_amount.data or loan.loan_amount
            loan.disbursed_amount = loan.approved_amount or loan.loan_amount
            loan.disbursement_date = form.disbursement_date.data
            loan.disbursement_method = form.disbursement_method.data
            loan.disbursement_reference = form.disbursement_reference.data
            loan.first_installment_date = form.first_installment_date.data
            if loan.loan_type != 'special_loan':
                loan.maturity_date = loan.disbursement_date + relativedelta(months=loan.duration_months) if loan.disbursement_date else None
            # Recalculate installment and totals from actual disbursed amount
            emi, total_payable = _calculate_loan_totals_for_principal(
                principal_amount=loan.disbursed_amount or loan.loan_amount,
                interest_rate=loan.interest_rate,
                loan_type=loan.loan_type,
                interest_type=loan.interest_type,
                duration_months=loan.duration_months,
                duration_weeks=loan.duration_weeks,
                duration_days=loan.duration_days,
            )
            loan.installment_amount = emi
            loan.total_payable = total_payable
            loan.paid_amount = Decimal(str(loan.paid_amount or 0)).quantize(Decimal('0.01'))
            loan.update_outstanding_amount()
            
            # Log activity
            log = ActivityLog(
                user_id=current_user.id,
                action='approve_loan',
                entity_type='loan',
                entity_id=loan.id,
                description=f'Approved loan: {loan.loan_number}',
                ip_address=request.remote_addr
            )
            flash('Loan approved successfully!', 'success')
        else:
            loan.status = 'rejected'
            loan.rejection_reason = form.rejection_reason.data
            
            # Log activity
            log = ActivityLog(
                user_id=current_user.id,
                action='reject_loan',
                entity_type='loan',
                entity_id=loan.id,
                description=f'Rejected loan: {loan.loan_number}',
                ip_address=request.remote_addr
            )
            flash('Loan rejected!', 'warning')
        
        db.session.add(log)
        db.session.commit()
        
        return redirect(url_for('loans.view_loan', id=id))
    
    # Pre-fill form defaults
    if not form.is_submitted():
        form.disbursement_date.data = datetime.now().date()
        # Calculate first installment date based on loan type
        if loan.duration_days:
            form.first_installment_date.data = datetime.now().date() + timedelta(days=1)
        elif loan.duration_weeks:
            form.first_installment_date.data = datetime.now().date() + timedelta(days=7)
        else:
            form.first_installment_date.data = datetime.now().date() + timedelta(days=30)
    
    return render_template('loans/approve.html',
                         title=f'Approve Loan: {loan.loan_number}',
                         form=form,
                         loan=loan)

@loans_bp.route('/<int:id>/approve-staff', methods=['GET', 'POST'])
@login_required
@permission_required('manage_loans')
def approve_loan_staff(id):
    """Staff approval (First stage)"""
    loan = Loan.query.get_or_404(id)
    
    # Check branch access
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id and loan.branch_id != current_branch_id:
            flash('Access denied: Loan not found in current branch.', 'danger')
            return redirect(url_for('loans.list_loans'))
    
    # Check if loan is in correct status
    if loan.status != 'pending':
        flash('Only pending loans can be approved by staff!', 'warning')
        return redirect(url_for('loans.view_loan', id=id))
    
    # Check if user is staff (not manager or admin)
    if current_user.role not in ['staff', 'loan_collector']:
        flash('Only staff members can perform first-stage approval!', 'warning')
        return redirect(url_for('loans.view_loan', id=id))
    
    form = StaffApprovalForm()
    
    if form.validate_on_submit():
        if form.approval_status.data == 'approve':
            loan.status = 'pending_manager_approval'
            loan.staff_approved_by = current_user.id
            loan.staff_approval_date = form.approval_date.data
            loan.staff_approval_notes = form.approval_notes.data
            
            # Log activity
            log = ActivityLog(
                user_id=current_user.id,
                action='staff_approve_loan',
                entity_type='loan',
                entity_id=loan.id,
                description=f'Staff approved loan: {loan.loan_number}',
                ip_address=request.remote_addr
            )
            flash('Loan approved by staff! Awaiting manager approval.', 'success')
        else:
            loan.status = 'rejected'
            loan.rejection_reason = form.rejection_reason.data
            
            # Log activity
            log = ActivityLog(
                user_id=current_user.id,
                action='staff_reject_loan',
                entity_type='loan',
                entity_id=loan.id,
                description=f'Staff rejected loan: {loan.loan_number}',
                ip_address=request.remote_addr
            )
            flash('Loan rejected by staff!', 'warning')
        
        db.session.add(log)
        db.session.commit()
        
        return redirect(url_for('loans.view_loan', id=id))
    
    # Pre-fill form
    form.approval_date.data = datetime.now().date()
    
    return render_template('loans/approve_staff.html',
                         title=f'Staff Approval: {loan.loan_number}',
                         form=form,
                         loan=loan)

@loans_bp.route('/<int:id>/approve-manager', methods=['GET', 'POST'])
@login_required
def approve_loan_manager(id):
    """Manager approval (Second stage)"""
    loan = Loan.query.get_or_404(id)
    
    # Check branch access
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id and loan.branch_id != current_branch_id:
            flash('Access denied: Loan not found in current branch.', 'danger')
            return redirect(url_for('loans.list_loans'))
    
    # Check if loan is in correct status
    if loan.status != 'pending_manager_approval':
        flash('Only loans approved by staff can be approved by manager!', 'warning')
        return redirect(url_for('loans.view_loan', id=id))
    
    # Check if user is manager
    if current_user.role not in ['manager', 'accountant']:
        flash('Only managers can perform second-stage approval!', 'warning')
        return redirect(url_for('loans.view_loan', id=id))
    
    form = ManagerApprovalForm()
    
    if form.validate_on_submit():
        if form.approval_status.data == 'approve':
            loan.status = 'initiated'
            loan.manager_approved_by = current_user.id
            loan.manager_approval_date = form.approval_date.data
            loan.manager_approval_notes = form.approval_notes.data
            
            # Log activity
            log = ActivityLog(
                user_id=current_user.id,
                action='manager_approve_loan',
                entity_type='loan',
                entity_id=loan.id,
                description=f'Manager approved loan: {loan.loan_number}',
                ip_address=request.remote_addr
            )
            flash('Loan approved by manager! Loan is now initiated. Awaiting admin approval for disbursement.', 'success')
        else:
            loan.status = 'rejected'
            loan.rejection_reason = form.rejection_reason.data
            
            # Log activity
            log = ActivityLog(
                user_id=current_user.id,
                action='manager_reject_loan',
                entity_type='loan',
                entity_id=loan.id,
                description=f'Manager rejected loan: {loan.loan_number}',
                ip_address=request.remote_addr
            )
            flash('Loan rejected by manager!', 'warning')
        
        db.session.add(log)
        db.session.commit()
        
        return redirect(url_for('loans.view_loan', id=id))
    
    # Pre-fill form
    form.approval_date.data = datetime.now().date()
    
    return render_template('loans/approve_manager.html',
                         title=f'Manager Approval: {loan.loan_number}',
                         form=form,
                         loan=loan)

@loans_bp.route('/<int:id>/approve-admin', methods=['GET', 'POST'])
@login_required
def approve_loan_admin(id):
    """Admin approval (Final stage - Disburse loan)"""
    loan = Loan.query.get_or_404(id)
    
    # Check branch access
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id and loan.branch_id != current_branch_id:
            flash('Access denied: Loan not found in current branch.', 'danger')
            return redirect(url_for('loans.list_loans'))
    
    # Check if loan is in correct status
    if loan.status != 'initiated':
        flash('Only initiated loans can be approved by admin!', 'warning')
        return redirect(url_for('loans.view_loan', id=id))
    
    is_admin_or_manager = current_user.role in ['admin', 'regional_manager']
    is_designated_approver = loan.final_approver_id is not None and current_user.id == loan.final_approver_id

    # Must be admin/regional_manager OR the designated final approver
    if not is_admin_or_manager and not is_designated_approver:
        flash('Only admins can perform final approval and disbursement!', 'warning')
        return redirect(url_for('loans.view_loan', id=id))

    # If admin/manager but a different person is designated, block them
    if is_admin_or_manager and loan.final_approver_id and current_user.id != loan.final_approver_id:
        designated = User.query.get(loan.final_approver_id)
        name = designated.full_name if designated else 'Unknown'
        flash(f'This loan can only be finally approved by the designated approver: {name}.', 'danger')
        return redirect(url_for('loans.view_loan', id=id))
    
    form = AdminApprovalForm()
    
    if form.validate_on_submit():
        if form.approval_status.data == 'approve':
            from decimal import Decimal
            loan.status = 'active'
            loan.admin_approved_by = current_user.id
            loan.admin_approval_date = form.approval_date.data
            loan.admin_approval_notes = form.approval_notes.data
            
            # Set approved_by for legacy compatibility
            loan.approved_by = current_user.id
            loan.approval_date = form.approval_date.data
            loan.approval_notes = form.approval_notes.data
            
            # Set disbursement details
            loan.approved_amount = form.approved_amount.data or loan.loan_amount
            loan.disbursed_amount = loan.approved_amount or loan.loan_amount
            loan.disbursement_date = form.disbursement_date.data
            loan.disbursement_method = form.disbursement_method.data
            loan.disbursement_reference = form.disbursement_reference.data
            loan.first_installment_date = form.first_installment_date.data
            
            # Calculate maturity date based on loan type (skip for special loans - already set)
            if loan.loan_type == 'special_loan':
                pass  # maturity_date already set from end_date during creation
            elif loan.duration_days:
                loan.maturity_date = loan.disbursement_date + timedelta(days=loan.duration_days) if loan.disbursement_date else None
            elif loan.duration_weeks:
                loan.maturity_date = loan.disbursement_date + timedelta(weeks=loan.duration_weeks) if loan.disbursement_date else None
            else:
                loan.maturity_date = loan.disbursement_date + relativedelta(months=loan.duration_months) if loan.disbursement_date else None
            
            # Recalculate installment and totals from actual disbursed amount
            emi, total_payable = _calculate_loan_totals_for_principal(
                principal_amount=loan.disbursed_amount or loan.loan_amount,
                interest_rate=loan.interest_rate,
                loan_type=loan.loan_type,
                interest_type=loan.interest_type,
                duration_months=loan.duration_months,
                duration_weeks=loan.duration_weeks,
                duration_days=loan.duration_days,
            )
            loan.installment_amount = emi
            loan.total_payable = total_payable
            loan.paid_amount = Decimal(str(loan.paid_amount or 0)).quantize(Decimal('0.01'))
            loan.update_outstanding_amount()
            
            # Log activity
            log = ActivityLog(
                user_id=current_user.id,
                action='admin_approve_loan',
                entity_type='loan',
                entity_id=loan.id,
                description=f'Admin approved and disbursed loan: {loan.loan_number}',
                ip_address=request.remote_addr
            )
            flash('Loan approved and disbursed successfully!', 'success')
        else:
            loan.status = 'rejected'
            loan.rejection_reason = form.rejection_reason.data
            
            # Log activity
            log = ActivityLog(
                user_id=current_user.id,
                action='admin_reject_loan',
                entity_type='loan',
                entity_id=loan.id,
                description=f'Admin rejected loan: {loan.loan_number}',
                ip_address=request.remote_addr
            )
            flash('Loan rejected by admin!', 'warning')
        
        db.session.add(log)
        db.session.commit()
        
        return redirect(url_for('loans.view_loan', id=id))
    
    # Pre-fill form
    form.approval_date.data = datetime.now().date()
    form.approved_amount.data = loan.loan_amount
    form.disbursement_date.data = datetime.now().date()
    
    # Calculate first installment date based on loan type
    if loan.duration_days:
        form.first_installment_date.data = datetime.now().date() + timedelta(days=1)
    elif loan.duration_weeks:
        form.first_installment_date.data = datetime.now().date() + timedelta(days=7)
    else:
        form.first_installment_date.data = datetime.now().date() + timedelta(days=30)
    
    return render_template('loans/approve_admin.html',
                         title=f'Admin Approval: {loan.loan_number}',
                         form=form,
                         loan=loan)

@loans_bp.route('/<int:id>/deactivate', methods=['GET', 'POST'])
@login_required
@permission_required('manage_loans')
def deactivate_loan(id):
    """Deactivate loan"""
    loan = Loan.query.get_or_404(id)
    
    # Check branch access
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id and loan.branch_id != current_branch_id:
            flash('Access denied: Loan not found in current branch.', 'danger')
            return redirect(url_for('loans.list_loans'))
    
    # Check if user has admin role for deactivation
    if current_user.role not in ['admin', 'regional_manager']:
        flash('Only administrators can deactivate loans.', 'danger')
        return redirect(url_for('loans.view_loan', id=id))
    
    # Check if loan can be deactivated
    if loan.status in ['completed', 'defaulted', 'rejected', 'deactivated']:
        flash('This loan cannot be deactivated.', 'danger')
        return redirect(url_for('loans.view_loan', id=id))
    
    form = LoanDeactivationForm()
    if form.validate_on_submit():
        if not form.confirm_deactivation.data:
            flash('Please confirm deactivation.', 'warning')
            return redirect(url_for('loans.deactivate_loan', id=id))
        
        # Deactivate the loan
        loan.status = 'deactivated'
        loan.deactivation_reason = form.deactivation_reason.data
        loan.deactivation_date = form.deactivation_date.data
        loan.deactivated_by = current_user.id
        
        # Log activity
        log = ActivityLog(
            user_id=current_user.id,
            action='deactivate_loan',
            entity_type='loan',
            entity_id=loan.id,
            description=f'Loan deactivated: {loan.loan_number} - Reason: {form.deactivation_reason.data}',
            ip_address=request.remote_addr
        )
        
        db.session.add(log)
        db.session.commit()
        
        flash('Loan deactivated successfully!', 'success')
        return redirect(url_for('loans.view_loan', id=id))
    
    return render_template('loans/deactivate.html',
                         title=f'Deactivate Loan: {loan.loan_number}',
                         form=form,
                         loan=loan)

def _process_payment(loan, payment_amount, payment_date, payment_method, reference_number, notes, penalty_amount=0):
    """Shared payment processor used by both form and quick-pay (keeps logic in one place)."""
    from decimal import Decimal, ROUND_HALF_UP

    payment_amount = Decimal(str(payment_amount)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    # Calculate current outstanding with accrued interest
    current_outstanding = loan.calculate_current_outstanding()
    accrued_interest = loan.calculate_accrued_interest()

    # Current outstanding principal (without accrued interest)
    disbursed = Decimal(str(loan.disbursed_amount or loan.loan_amount))
    paid_principal = loan.get_total_paid_principal()
    outstanding_principal = disbursed - paid_principal

    # Check if this is a full payment or overpayment
    is_full_payment = payment_amount >= current_outstanding or abs(payment_amount - current_outstanding) <= Decimal('0.05')

    # Calculate interest and principal splits based on loan type
    if loan.interest_type == 'flat':
        # For flat interest loans, calculate fixed interest per payment
        # Determine number of installments and period based on frequency
        if loan.installment_frequency == 'monthly':
            num_installments = loan.duration_months or 0
            period = Decimal('1') / Decimal('12')
        elif loan.installment_frequency == 'weekly':
            num_installments = loan.duration_weeks or 0
            period = Decimal('7') / Decimal('365')
        elif loan.installment_frequency == 'daily':
            num_installments = loan.duration_days or 0
            period = Decimal('1') / Decimal('365')
        else:
            num_installments = loan.duration_months or 0
            period = Decimal('1') / Decimal('12')

        time_in_years = Decimal(str(num_installments)) * period
        total_interest = (disbursed * Decimal(str(loan.interest_rate)) * time_in_years) / Decimal('100')
        installment_interest = total_interest / Decimal(str(num_installments)) if num_installments > 0 else Decimal('0')

        # For flat loans, determine if this is truly a full settlement (paying off entire remaining balance)
        total_paid = Decimal(str(loan.paid_amount or 0))
        total_payable = disbursed + total_interest
        remaining_total = total_payable - total_paid
        is_full_settlement = payment_amount >= remaining_total or abs(payment_amount - remaining_total) <= Decimal('0.05')

        if is_full_settlement:
            # Full settlement - pay all remaining interest and principal
            total_paid_interest = loan.get_total_paid_interest()
            remaining_interest = total_interest - Decimal(str(total_paid_interest))
            interest_amount = min(payment_amount, remaining_interest)
            principal_amount = payment_amount - interest_amount
        else:
            # Regular installment payment - use fixed interest per payment
            if loan.installment_amount and loan.installment_amount > 0 and abs(payment_amount - Decimal(str(loan.installment_amount))) <= Decimal('0.05'):
                # This is a regular installment payment
                interest_amount = installment_interest
                principal_amount = payment_amount - interest_amount
            else:
                # Partial or different amount - calculate proportionally
                payment_ratio = payment_amount / Decimal(str(loan.installment_amount)) if loan.installment_amount else Decimal('1')
                interest_amount = installment_interest * payment_ratio
                principal_amount = payment_amount - interest_amount
    else:
        # For reducing balance loans (original logic)
        if is_full_payment:
            # For full payments, pay off accrued interest first, then principal
            interest_amount = min(payment_amount, accrued_interest)
            principal_amount = min(payment_amount - interest_amount, outstanding_principal)
            
            # Any remaining amount goes to interest (overpayment case)
            remaining = payment_amount - principal_amount - interest_amount
            if remaining > 0:
                interest_amount += remaining
        else:
            # For partial payments, split based on loan terms
            if accrued_interest > 0:
                # Pay accrued interest first if any exists
                interest_amount = min(payment_amount, accrued_interest)
                principal_amount = payment_amount - interest_amount
            else:
                # Standard EMI split - calculate proportion
                monthly_interest = outstanding_principal * (Decimal(str(loan.interest_rate)) / Decimal('1200'))
                if payment_amount >= monthly_interest:
                    interest_amount = monthly_interest
                    principal_amount = payment_amount - interest_amount
                else:
                    # If payment is less than monthly interest, all goes to interest
                    interest_amount = payment_amount
                    principal_amount = Decimal('0')
    # Ensure no negative amounts
    interest_amount = max(Decimal('0'), interest_amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
    principal_amount = max(Decimal('0'), principal_amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
    
    # Create payment record
    payment = LoanPayment(
        loan_id=loan.id,
        payment_date=payment_date,
        payment_amount=float(payment_amount),
        principal_amount=float(principal_amount),
        interest_amount=float(interest_amount),
        penalty_amount=float(penalty_amount or 0),
        payment_method=payment_method,
        reference_number=reference_number,
        receipt_number=generate_receipt_number(),
        notes=notes,
        collected_by=current_user.id
    )
    
    db.session.add(payment)
    
    # Update loan amounts - recalculate outstanding based on new payment
    loan.paid_amount = (Decimal(str(loan.paid_amount or 0)) + payment_amount).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    loan.update_outstanding_amount()  # Use our new method to calculate current outstanding
    
    # Calculate advance balance from schedule allocation (not due-so-far shortcut)
    schedule = loan.generate_payment_schedule()
    loan.advance_balance = loan.calculate_available_advance_balance(schedule=schedule)
    
    # Set balance_after for the payment record
    payment.balance_after = float(loan.outstanding_amount or 0)
    
    # Check if loan is fully paid
    if loan.calculate_current_outstanding() <= Decimal('0.02'):  # Allow for small rounding differences
        loan.status = 'completed'
        loan.outstanding_amount = Decimal('0')
        loan.advance_balance = Decimal('0')
        loan.closing_date = payment_date  # Set closing date to payment date
        
    # Log activity
    log = ActivityLog(
        user_id=current_user.id,
        action='loan_payment',
        entity_type='loan',
        entity_id=loan.id,
        description=f'Payment of {payment_amount} for loan {loan.loan_number}',
        ip_address=request.remote_addr
    )
    db.session.add(log)
    
    db.session.commit()
    return payment


@loans_bp.route('/<int:id>/payment', methods=['GET', 'POST'])
@login_required
@permission_required('collect_payments')
def add_payment(id):
    """Add loan payment"""
    from decimal import Decimal

    loan = Loan.query.get_or_404(id)

    # Check branch access
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id and loan.branch_id != current_branch_id:
            flash('Access denied: Loan not found in current branch.', 'danger')
            return redirect(url_for('loans.list_loans'))

    if loan.status not in ['active', 'disbursed']:
        flash('Cannot add payment for this loan! Loan must be active.', 'warning')
        return redirect(url_for('loans.view_loan', id=id))

    form = LoanPaymentForm()
    advance_breakdown = _get_installment_advance_breakdown(loan)

    if form.validate_on_submit():
        use_advance_credit = request.form.get('use_advance_credit') == '1'
        payment_amount, auto_applied_advance = _resolve_payment_amount_with_optional_advance(
            posted_amount=form.payment_amount.data,
            use_advance_credit=use_advance_credit,
            installment_amount=advance_breakdown['installment_amount'],
            advance_to_apply=advance_breakdown['advance_to_apply'],
        )

        if auto_applied_advance:
            flash(
                f'Advance credit deducted: {advance_breakdown["advance_to_apply"]}. '
                f'Cash collected set to {payment_amount}.',
                'info',
            )

        _process_payment(
            loan=loan,
            payment_amount=payment_amount,
            payment_date=form.payment_date.data,
            payment_method=form.payment_method.data,
            reference_number=form.reference_number.data,
            notes=form.notes.data,
            penalty_amount=form.penalty_amount.data,
        )

        flash(f'Payment of {payment_amount} recorded successfully!', 'success')
        return redirect(url_for('loans.view_loan', id=id))
    
    # Set default values only for GET requests and calculate current amounts
    if request.method == 'GET':
        # Default to installment after advance (if available), while UI allows toggling.
        if advance_breakdown['advance_to_apply'] > Decimal('0.00'):
            form.payment_amount.data = float(advance_breakdown['deducted_installment_amount'])
        else:
            form.payment_amount.data = float(advance_breakdown['installment_amount'])
        form.payment_date.data = datetime.now().date()
    
    # Get current outstanding amount with accrued interest for display
    current_outstanding = loan.calculate_current_outstanding()
    accrued_interest = loan.calculate_accrued_interest()
    
    # Get arrears details for display
    arrears_details = loan.get_arrears_details()
    advance_balance = float(advance_breakdown['advance_balance'])
    
    # Update stored outstanding amount to match current calculation
    loan.update_outstanding_amount()
    db.session.commit()
    
    return render_template('loans/payment.html',
                         title=f'Add Payment: {loan.loan_number}',
                         form=form,
                         loan=loan,
                         current_outstanding=current_outstanding,
                         accrued_interest=accrued_interest,
                         arrears_details=arrears_details,
                         advance_balance=advance_balance,
                         installment_amount=float(advance_breakdown['installment_amount']),
                         advance_to_apply=float(advance_breakdown['advance_to_apply']),
                         deducted_installment_amount=float(advance_breakdown['deducted_installment_amount']),
                         use_advance_credit_default=advance_breakdown['advance_to_apply'] > Decimal('0.00'))

@loans_bp.route('/search')
@login_required
@permission_required('manage_loans')
def search_loans():
    """Search loans page"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    loan_type = request.args.get('loan_type', '', type=str)
    status = request.args.get('status', '', type=str)
    interest_type = request.args.get('interest_type', '', type=str)
    min_amount = request.args.get('min_amount', '', type=str)
    max_amount = request.args.get('max_amount', '', type=str)
    
    loans = None
    searched = False
    
    if search or loan_type or status or interest_type or min_amount or max_amount:
        searched = True
        query = Loan.query
        
        # Filter by current branch if needed
        if should_filter_by_branch():
            current_branch_id = get_current_branch_id()
            if current_branch_id:
                query = query.filter_by(branch_id=current_branch_id)
        
        if search:
            query = query.join(Customer).filter(
                db.or_(
                    Loan.loan_number.ilike(f'%{search}%'),
                    Customer.full_name.ilike(f'%{search}%'),
                    Customer.customer_id.ilike(f'%{search}%')
                )
            )
        
        if loan_type:
            query = query.filter_by(loan_type=loan_type)
        
        if status:
            query = query.filter_by(status=status)
        
        if interest_type:
            query = query.filter_by(interest_type=interest_type)
        
        if min_amount:
            try:
                query = query.filter(Loan.loan_amount >= float(min_amount))
            except ValueError:
                pass
        
        if max_amount:
            try:
                query = query.filter(Loan.loan_amount <= float(max_amount))
            except ValueError:
                pass
        
        loans = query.order_by(Loan.created_at.desc()).paginate(
            page=page, per_page=current_app.config['ITEMS_PER_PAGE'], error_out=False
        )
    
    return render_template('loans/search.html',
                         title='Search Loans',
                         loans=loans,
                         search=search,
                         loan_type=loan_type,
                         status=status,
                         interest_type=interest_type,
                         min_amount=min_amount,
                         max_amount=max_amount,
                         searched=searched)

# API endpoint for fetching guarantors
@loans_bp.route('/api/guarantors')
@login_required
@permission_required('manage_loans')
def get_guarantors():
    """Get KYC approved guarantors and family guarantors"""
    # Get the member ID to exclude (loan borrower cannot be their own guarantor)
    exclude_customer_id = request.args.get('exclude_customer_id', type=int)
    
    # Query customers who are guarantors or family guarantors and have KYC verified
    query = Customer.query.filter(
        db.or_(
            Customer.customer_type.like('%guarantor%'),
            Customer.customer_type.like('%family_guarantor%')
        ),
        Customer.kyc_verified == True
    )
    
    # Exclude the selected loan member
    if exclude_customer_id:
        query = query.filter(Customer.id != exclude_customer_id)
    
    # Filter by branch if needed
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id:
            query = query.filter_by(branch_id=current_branch_id)
    
    guarantors = query.order_by(Customer.full_name).all()
    
    # Format guarantor data
    guarantor_list = []
    for guarantor in guarantors:
        guarantor_list.append({
            'id': guarantor.id,
            'full_name': guarantor.full_name,
            'nic_number': guarantor.nic_number,
            'phone_primary': guarantor.phone_primary,
            'customer_type': guarantor.customer_type,
            'customer_type_display': guarantor.customer_type_display,
            'address_line1': guarantor.address_line1,
            'address_line2': guarantor.address_line2 or '',
            'city': guarantor.city,
            'district': guarantor.district,
            'email': guarantor.email or ''
        })
    
    return jsonify({
        'success': True,
        'guarantors': guarantor_list
    })


# API endpoint for searching customers
@loans_bp.route('/api/search-customers')
@login_required
@permission_required('manage_loans')
def search_customers():
    """Search customers for loan assignment"""
    search_term = request.args.get('q', '', type=str).strip()
    loan_type = request.args.get('loan_type', '', type=str)
    
    if not search_term or len(search_term) < 2:
        return jsonify({'success': False, 'customers': []})
    
    # Staff Loan search is driven by active system users.
    if loan_type == 'staff_loan':
        user_query = User.query.filter(User.is_active == True)

        if should_filter_by_branch():
            current_branch_id = get_current_branch_id()
            if current_branch_id:
                user_query = user_query.filter(User.branch_id == current_branch_id)

        users = user_query.filter(
            db.or_(
                User.full_name.ilike(f'%{search_term}%'),
                User.username.ilike(f'%{search_term}%'),
                User.email.ilike(f'%{search_term}%'),
                User.nic_number.ilike(f'%{search_term}%')
            )
        ).order_by(User.full_name).limit(10).all()

        user_nics = [u.nic_number for u in users if u.nic_number]
        customer_by_nic = {}
        if user_nics:
            linked_customers = Customer.query.filter(Customer.nic_number.in_(user_nics)).all()
            customer_by_nic = {c.nic_number: c for c in linked_customers}

        customer_list = []
        for user in users:
            linked_customer = customer_by_nic.get(user.nic_number)
            if linked_customer and linked_customer.status != 'active':
                continue
            has_active_member = linked_customer and linked_customer.status == 'active'
            selection_id = linked_customer.id if has_active_member else -user.id

            customer_list.append({
                'id': selection_id,
                'text': f'{user.full_name} ({user.username}) - NIC: {user.nic_number}',
                'customer_id': linked_customer.customer_id if has_active_member else f'USER-{user.id}',
                'full_name': user.full_name,
                'nic_number': user.nic_number,
                'phone_primary': user.phone
            })
    else:
        query = Customer.query.filter(
            Customer.status == 'active',
            Customer.kyc_verified == True
        )

        # Filter by current branch if needed
        if should_filter_by_branch():
            current_branch_id = get_current_branch_id()
            if current_branch_id:
                query = query.filter_by(branch_id=current_branch_id)

        # Search by name, customer ID, or NIC number
        customers = query.filter(
            db.or_(
                Customer.full_name.ilike(f'%{search_term}%'),
                Customer.customer_id.ilike(f'%{search_term}%'),
                Customer.nic_number.ilike(f'%{search_term}%')
            )
        ).order_by(Customer.full_name).limit(10).all()

        # Format customer data
        customer_list = []
        for customer in customers:
            customer_list.append({
                'id': customer.id,
                'text': f'{customer.customer_id} - {customer.full_name} ({customer.nic_number})',
                'customer_id': customer.customer_id,
                'full_name': customer.full_name,
                'nic_number': customer.nic_number,
                'phone_primary': customer.phone_primary
            })
    
    return jsonify({
        'success': True,
        'customers': customer_list
    })


@loans_bp.route('/api/search-guarantors')
@login_required
@permission_required('manage_loans')
def search_guarantors():
    """Search guarantors with loan-type-specific eligibility rules."""
    search_term = request.args.get('q', '', type=str).strip()
    exclude_customer_id = request.args.get('exclude_customer_id', type=int)
    loan_type = request.args.get('loan_type', '', type=str)
    
    if not search_term or len(search_term) < 2:
        return jsonify({'success': False, 'guarantors': []})
    
    if loan_type == 'staff_loan':
        user_query = User.query.filter(User.is_active == True)

        if should_filter_by_branch():
            current_branch_id = get_current_branch_id()
            if current_branch_id:
                user_query = user_query.filter(User.branch_id == current_branch_id)

        exclude_user_id = None
        if exclude_customer_id:
            if exclude_customer_id < 0:
                exclude_user_id = abs(exclude_customer_id)
            else:
                borrower_customer = Customer.query.get(exclude_customer_id)
                if borrower_customer and borrower_customer.nic_number:
                    borrower_user = User.query.filter(
                        User.nic_number == borrower_customer.nic_number,
                        User.is_active == True
                    ).first()
                    if borrower_user:
                        exclude_user_id = borrower_user.id

        if exclude_user_id:
            user_query = user_query.filter(User.id != exclude_user_id)

        users = user_query.filter(
            db.or_(
                User.full_name.ilike(f'%{search_term}%'),
                User.username.ilike(f'%{search_term}%'),
                User.email.ilike(f'%{search_term}%'),
                User.nic_number.ilike(f'%{search_term}%')
            )
        ).order_by(User.full_name).limit(10).all()

        user_nics = [u.nic_number for u in users if u.nic_number]
        customer_by_nic = {}
        if user_nics:
            linked_customers = Customer.query.filter(Customer.nic_number.in_(user_nics)).all()
            customer_by_nic = {c.nic_number: c for c in linked_customers}

        guarantor_list = []
        for user in users:
            linked_customer = customer_by_nic.get(user.nic_number)
            if linked_customer and linked_customer.status != 'active':
                continue

            has_active_member = linked_customer is not None
            selection_id = linked_customer.id if has_active_member else -user.id

            guarantor_list.append({
                'id': selection_id,
                'text': f'{user.full_name} ({user.username}) - NIC: {user.nic_number}',
                'customer_id': linked_customer.customer_id if has_active_member else f'USER-{user.id}',
                'full_name': user.full_name,
                'nic_number': user.nic_number,
                'phone_primary': user.phone,
                'customer_type_display': 'Staff User',
                'address_line1': linked_customer.address_line1 if has_active_member else 'From Settings Users',
                'address_line2': (linked_customer.address_line2 if has_active_member and linked_customer.address_line2 else ''),
                'city': linked_customer.city if has_active_member else '',
                'district': linked_customer.district if has_active_member else '',
                'email': user.email or '',
                'eligibility_text': 'Active Settings User'
            })
    else:
        # Non-staff loans use KYC-approved guarantor member records only.
        query = Customer.query.filter(
            db.or_(
                Customer.customer_type.like('%guarantor%'),
                Customer.customer_type.like('%family_guarantor%')
            ),
            Customer.kyc_verified == True,
            Customer.status == 'active'
        )

        if exclude_customer_id:
            query = query.filter(Customer.id != exclude_customer_id)

        if should_filter_by_branch():
            current_branch_id = get_current_branch_id()
            if current_branch_id:
                query = query.filter_by(branch_id=current_branch_id)

        guarantors = query.filter(
            db.or_(
                Customer.full_name.ilike(f'%{search_term}%'),
                Customer.customer_id.ilike(f'%{search_term}%'),
                Customer.nic_number.ilike(f'%{search_term}%')
            )
        ).order_by(Customer.full_name).limit(10).all()

        guarantor_list = []
        for guarantor in guarantors:
            guarantor_list.append({
                'id': guarantor.id,
                'text': f'{guarantor.customer_id} - {guarantor.full_name} ({guarantor.nic_number})',
                'customer_id': guarantor.customer_id,
                'full_name': guarantor.full_name,
                'nic_number': guarantor.nic_number,
                'phone_primary': guarantor.phone_primary,
                'customer_type_display': guarantor.customer_type_display,
                'address_line1': guarantor.address_line1,
                'address_line2': guarantor.address_line2 or '',
                'city': guarantor.city,
                'district': guarantor.district,
                'email': guarantor.email or '',
                'eligibility_text': 'KYC Verified'
            })
    
    return jsonify({
        'success': True,
        'guarantors': guarantor_list
    })


@loans_bp.route('/payment/<int:payment_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_payment(payment_id):
    """Edit an existing loan payment (Admin only)"""
    payment = LoanPayment.query.get_or_404(payment_id)
    loan = payment.loan

    # Check branch access
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id and loan.branch_id != current_branch_id:
            flash('Access denied: Payment not found in current branch.', 'danger')
            return redirect(url_for('loans.list_loans'))

    form = EditPaymentForm()

    if request.method == 'GET':
        form.payment_date.data = payment.payment_date
        form.payment_amount.data = payment.payment_amount
        form.principal_amount.data = payment.principal_amount
        form.interest_amount.data = payment.interest_amount
        form.penalty_amount.data = payment.penalty_amount or 0
        form.balance_after.data = payment.balance_after
        form.payment_method.data = payment.payment_method
        form.reference_number.data = payment.reference_number
        form.notes.data = payment.notes

    if form.validate_on_submit():
        from decimal import Decimal, ROUND_HALF_UP

        old_amount = Decimal(str(payment.payment_amount))
        new_amount = Decimal(str(form.payment_amount.data)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        diff = new_amount - old_amount

        payment.payment_date = form.payment_date.data
        payment.payment_amount = float(new_amount)
        payment.principal_amount = float(form.principal_amount.data) if form.principal_amount.data is not None else payment.principal_amount
        payment.interest_amount = float(form.interest_amount.data) if form.interest_amount.data is not None else payment.interest_amount
        payment.penalty_amount = float(form.penalty_amount.data or 0)
        payment.balance_after = float(form.balance_after.data) if form.balance_after.data is not None else payment.balance_after
        payment.payment_method = form.payment_method.data
        payment.reference_number = form.reference_number.data
        payment.notes = form.notes.data

        # Adjust loan paid_amount by the difference
        if diff != 0:
            loan.paid_amount = (Decimal(str(loan.paid_amount or 0)) + diff).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            _refresh_loan_financial_state(loan)

        # Log activity
        log = ActivityLog(
            user_id=current_user.id,
            action='edit_payment',
            entity_type='loan',
            entity_id=loan.id,
            description=f'Edited payment {payment.receipt_number} for loan {loan.loan_number}',
            ip_address=request.remote_addr
        )
        db.session.add(log)
        db.session.commit()

        flash(f'Payment {payment.receipt_number} updated successfully!', 'success')
        return redirect(url_for('loans.view_loan', id=loan.id))

    return render_template('loans/edit_payment.html',
                           title=f'Edit Payment: {payment.receipt_number}',
                           form=form,
                           payment=payment,
                           loan=loan)


@loans_bp.route('/payment/<int:payment_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_payment(payment_id):
    """Delete a loan payment (Admin only)"""
    payment = LoanPayment.query.get_or_404(payment_id)
    loan = payment.loan

    # Check branch access
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id and loan.branch_id != current_branch_id:
            flash('Access denied: Payment not found in current branch.', 'danger')
            return redirect(url_for('loans.list_loans'))

    from decimal import Decimal, ROUND_HALF_UP
    receipt_number = payment.receipt_number
    amount = Decimal(str(payment.payment_amount))

    # Reverse the payment amount from loan's paid_amount
    loan.paid_amount = (Decimal(str(loan.paid_amount or 0)) - amount).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    if loan.paid_amount < 0:
        loan.paid_amount = Decimal('0')
    _refresh_loan_financial_state(loan)

    db.session.delete(payment)

    log = ActivityLog(
        user_id=current_user.id,
        action='delete_payment',
        entity_type='loan',
        entity_id=loan.id,
        description=f'Deleted payment {receipt_number} (Rs. {amount}) from loan {loan.loan_number}',
        ip_address=request.remote_addr
    )
    db.session.add(log)
    db.session.commit()

    flash(f'Payment {receipt_number} deleted successfully. Loan balance has been updated.', 'success')
    return redirect(url_for('loans.view_loan', id=loan.id))


@loans_bp.route('/receipt-entry')
@login_required
@permission_required('collect_payments')
def receipt_entry():
    """Receipt entry page with weekly, daily, monthly, staff, and special loan tables."""
    referrer = request.args.get('collector', type=int)
    
    # Get weekly loans (type1_9weeks and type4_micro)
    weekly_loans_query = Loan.query.filter(
        Loan.loan_type.in_(['type1_9weeks', 'type4_micro']),
        Loan.status == 'active'
    )
    
    # Get daily loans (54_daily and type4_daily)
    daily_loans_query = Loan.query.filter(
        Loan.loan_type.in_(['54_daily', 'type4_daily']),
        Loan.status == 'active'
    )
    
    # Get monthly loans (monthly_loan)
    monthly_loans_query = Loan.query.filter(
        Loan.loan_type.in_(['monthly_loan']),
        Loan.status == 'active'
    )

    # Get staff loans (staff_loan)
    staff_loans_query = Loan.query.filter(
        Loan.loan_type.in_(['staff_loan']),
        Loan.status == 'active'
    )
    
    # Get special loans
    special_loans_query = Loan.query.filter(
        Loan.loan_type.in_(['special_loan']),
        Loan.status == 'active'
    )
    
    # Filter by branch if needed
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id:
            weekly_loans_query = weekly_loans_query.filter_by(branch_id=current_branch_id)
            daily_loans_query = daily_loans_query.filter_by(branch_id=current_branch_id)
            monthly_loans_query = monthly_loans_query.filter_by(branch_id=current_branch_id)
            staff_loans_query = staff_loans_query.filter_by(branch_id=current_branch_id)
            special_loans_query = special_loans_query.filter_by(branch_id=current_branch_id)
    
    # Filter by referrer if specified
    if referrer:
        weekly_loans_query = weekly_loans_query.filter(Loan.referred_by == referrer)
        daily_loans_query = daily_loans_query.filter(Loan.referred_by == referrer)
        monthly_loans_query = monthly_loans_query.filter(Loan.referred_by == referrer)
        staff_loans_query = staff_loans_query.filter(Loan.referred_by == referrer)
        special_loans_query = special_loans_query.filter(Loan.referred_by == referrer)
    
    weekly_loans = weekly_loans_query.order_by(Loan.created_at.desc()).all()
    daily_loans = daily_loans_query.order_by(Loan.created_at.desc()).all()
    monthly_loans = monthly_loans_query.order_by(Loan.created_at.desc()).all()
    staff_loans = staff_loans_query.order_by(Loan.created_at.desc()).all()
    special_loans = special_loans_query.order_by(Loan.created_at.desc()).all()
    
    # Get recent payments for each loan with collector info
    weekly_payments = []
    for loan in weekly_loans:
        recent_payments = loan.payments.order_by(LoanPayment.payment_date.desc()).limit(5).all()
        weekly_payments.append({
            'loan': loan,
            'recent_payments': recent_payments,
            'recommended_amount': loan.get_next_installment_amount()
        })
    
    daily_payments = []
    for loan in daily_loans:
        recent_payments = loan.payments.order_by(LoanPayment.payment_date.desc()).limit(5).all()
        daily_payments.append({
            'loan': loan,
            'recent_payments': recent_payments,
            'recommended_amount': loan.get_next_installment_amount()
        })
    
    monthly_payments = []
    for loan in monthly_loans:
        recent_payments = loan.payments.order_by(LoanPayment.payment_date.desc()).limit(5).all()
        monthly_payments.append({
            'loan': loan,
            'recent_payments': recent_payments,
            'recommended_amount': loan.get_next_installment_amount()
        })

    staff_payments = []
    for loan in staff_loans:
        recent_payments = loan.payments.order_by(LoanPayment.payment_date.desc()).limit(5).all()
        staff_payments.append({
            'loan': loan,
            'recent_payments': recent_payments,
            'recommended_amount': loan.get_next_installment_amount()
        })
    
    special_payments = []
    for loan in special_loans:
        recent_payments = loan.payments.order_by(LoanPayment.payment_date.desc()).limit(5).all()
        special_payments.append({
            'loan': loan,
            'recent_payments': recent_payments,
            'recommended_amount': loan.get_next_installment_amount()
        })
    
    # Get all payments for payment history
    all_payments_query = LoanPayment.query.join(Loan)
    
    # Filter by branch if needed
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id:
            all_payments_query = all_payments_query.filter(Loan.branch_id == current_branch_id)
    
    # Filter by referrer if specified
    if referrer:
        all_payments_query = all_payments_query.filter(Loan.referred_by == referrer)
    
    all_payments = all_payments_query.order_by(LoanPayment.payment_date.desc()).limit(100).all()
    
    # Get users for referrer dropdown
    users_query = User.query.filter_by(is_active=True)
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id:
            users_query = users_query.filter_by(branch_id=current_branch_id)
    users = users_query.order_by(User.full_name).all()
    
    return render_template('loans/receipt_entry.html',
                         title='Receipt Entry',
                         weekly_payments=weekly_payments,
                         daily_payments=daily_payments,
                         monthly_payments=monthly_payments,
                         staff_payments=staff_payments,
                         special_payments=special_payments,
                         all_payments=all_payments,
                         users=users,
                         collector=referrer)


@loans_bp.route('/<int:id>/quick-pay', methods=['POST'])
@login_required
@permission_required('collect_payments')
def quick_pay(id):
    """Quick cash payment from receipt-entry tables."""
    loan = Loan.query.get_or_404(id)

    # Check branch access
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id and loan.branch_id != current_branch_id:
            return jsonify({'success': False, 'message': 'Access denied for this branch'}), 403

    if loan.status not in ['active', 'disbursed']:
        return jsonify({'success': False, 'message': 'Loan is not active'}), 400

    data = request.get_json() or {}
    amount = data.get('amount')
    try:
        amount_value = float(amount)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': 'Invalid amount'}), 400

    if amount_value <= 0:
        return jsonify({'success': False, 'message': 'Amount must be greater than zero'}), 400

    payment_date = datetime.now().date()

    try:
        _process_payment(
            loan=loan,
            payment_amount=amount_value,
            payment_date=payment_date,
            payment_method='cash',
            reference_number=None,
            notes='Quick pay (receipt entry)',
            penalty_amount=0,
        )
        return jsonify({'success': True, 'message': 'Payment recorded successfully'}), 200
    except Exception as exc:  # noqa: BLE001
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Failed to record payment: {exc}'}), 500


@loans_bp.route('/receipt-entry/export/<loan_frequency>')
@login_required
@permission_required('collect_payments')
def receipt_entry_export(loan_frequency):
    """Export weekly/daily/monthly/staff/special loans to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    frequency_map = {
        'weekly': (['type1_9weeks', 'type4_micro'], 'Weekly Loans'),
        'daily': (['54_daily', 'type4_daily'], 'Daily Loans'),
        'monthly': (['monthly_loan'], 'Monthly Loans'),
        'staff': (['staff_loan'], 'Staff Loans'),
        'special': (['special_loan'], 'Special Loans'),
    }
    if loan_frequency not in frequency_map:
        from flask import abort
        abort(404)

    loan_types, sheet_title = frequency_map[loan_frequency]
    referrer = request.args.get('collector', type=int)

    query = Loan.query.filter(
        Loan.loan_type.in_(loan_types),
        Loan.status == 'active'
    )
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id:
            query = query.filter_by(branch_id=current_branch_id)
    if referrer:
        query = query.filter(Loan.referred_by == referrer)

    loans = query.order_by(Loan.created_at.desc()).all()

    # Build schedules and collect all unique due dates
    loan_schedules = {}
    for loan in loans:
        schedule = loan.generate_payment_schedule()
        loan_schedules[loan.id] = {
            'num_arrears': sum(1 for inst in schedule if inst['status'] == 'overdue'),
        }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Match Repayment.pdf header exactly (including spelling)
    weekdays = ['Monday', 'Tuesday', 'Wednessday', 'Thursday', 'Friday', 'Saturday']
    fixed_headers = ['Loan Number', 'Customer', 'Phone', 'loan Amount', 'Installment', 'Arreas']
    n_fixed = len(fixed_headers)
    # Each weekday takes 1 column + Signature at the end
    total_cols = n_fixed + len(weekdays) + 1

    header_fill  = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font  = Font(bold=True, color='FFFFFF', size=9)
    sub_fill     = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    sub_font     = Font(bold=True, color='FFFFFF', size=8)
    thin_border  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_cell(cell, fill, font, h_align='center'):
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal=h_align, vertical='center', wrap_text=True)
        cell.border = thin_border

    # --- Row 1: empty header row (blank tabs above weekday names, as in Repayment.pdf) ---
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=1, column=c, value='')
        # Keep borders so the "tabs" are visible
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Row 2: actual header row (matches Repayment.pdf) ---
    col_idx = 1
    for name in fixed_headers:
        cell = ws.cell(row=2, column=col_idx, value=name)
        style_cell(cell, header_fill, header_font)
        col_idx += 1

    for day in weekdays:
        cell = ws.cell(row=2, column=col_idx, value=day)
        style_cell(cell, header_fill, header_font)
        col_idx += 1

    sig_col = total_cols
    cell = ws.cell(row=2, column=sig_col, value='Signature')
    style_cell(cell, header_fill, header_font)

    # --- Data rows (start at row 3) ---
    data_font = Font(size=9)
    for loan in loans:
        sched_data = loan_schedules.get(loan.id, {})
        num_arrears = sched_data.get('num_arrears', 0)
        customer = loan.customer

        row_values = [
            loan.loan_number,
            customer.full_name if customer else 'N/A',
            customer.phone_primary if customer else 'N/A',
            float(loan.loan_amount) if loan.loan_amount else 0,
            float(loan.installment_amount) if loan.installment_amount else 0,
            num_arrears,
        ]

        # 6 weekday blank columns + 1 blank Signature
        row_values += [''] * len(weekdays) + ['']

        ws.append(row_values)
        row_num = ws.max_row
        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    # --- TOTAL row ---
    total_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    total_font = Font(bold=True, size=9)
    data_start_row = 3
    data_end_row   = ws.max_row
    total_row = []

    for col_idx in range(1, total_cols + 1):
        if col_idx == 1:
            total_row.append('TOTAL')
        # Sum numeric columns: loan Amount (4), Installment (5), Arreas (6)
        elif col_idx in (4, 5, 6):
            col_letter = get_column_letter(col_idx)
            total_row.append(
                f'=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})'
                if data_end_row >= data_start_row else 0
            )
        else:
            total_row.append('')

    ws.append(total_row)
    total_row_num = ws.max_row
    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(row=total_row_num, column=col_idx)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left')

    # --- Column widths ---
    fixed_widths = [14, 22, 15, 14, 14, 10]  # Loan Number, Customer, Phone, loan Amount, Installment, Arreas
    for i, w in enumerate(fixed_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # Weekday columns (1 column each)
    for day_idx in range(len(weekdays)):
        day_col = n_fixed + day_idx + 1
        ws.column_dimensions[get_column_letter(day_col)].width = 12

    # Signature column
    sig_col = total_cols
    ws.column_dimensions[get_column_letter(sig_col)].width = 16

    ws.row_dimensions[1].height = 12
    ws.row_dimensions[2].height = 20

    # Freeze above data rows and keep fixed columns visible
    ws.freeze_panes = ws.cell(row=3, column=n_fixed + 1)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={loan_frequency}_loans_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return response


@loans_bp.route('/receipt-entry/pdf/<loan_frequency>')
@login_required
@permission_required('collect_payments')
def receipt_entry_pdf(loan_frequency):
    """Export receipt entry loan list to landscape PDF"""
    import html as _html
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

    def esc(value):
        """Escape a value for safe use inside a ReportLab Paragraph XML string."""
        return _html.escape(str(value) if value is not None else '')

    frequency_map = {
        'weekly': (['type1_9weeks', 'type4_micro'], 'Weekly Loans'),
        'daily': (['54_daily', 'type4_daily'], 'Daily Loans'),
        'monthly': (['monthly_loan'], 'Monthly Loans'),
        'staff': (['staff_loan'], 'Staff Loans'),
        'special': (['special_loan'], 'Special Loans'),
    }
    if loan_frequency not in frequency_map:
        from flask import abort
        abort(404)

    loan_types, title = frequency_map[loan_frequency]
    referrer = request.args.get('collector', type=int)

    query = Loan.query.filter(
        Loan.loan_type.in_(loan_types),
        Loan.status == 'active'
    )
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id:
            query = query.filter_by(branch_id=current_branch_id)
    if referrer:
        query = query.filter(Loan.referred_by == referrer)

    loans = query.order_by(Loan.created_at.desc()).all()

    # --- Build PDF ---
    buf = io.BytesIO()
    page_w, page_h = landscape(A4)
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'PDFTitle', parent=styles['Heading2'], fontSize=11, alignment=TA_CENTER, spaceAfter=4 * mm
    )
    cell_style = ParagraphStyle(
        'Cell', parent=styles['Normal'], fontSize=6.5, leading=8
    )
    cell_bold = ParagraphStyle(
        'CellBold', parent=cell_style, fontName='Helvetica-Bold'
    )
    cell_right = ParagraphStyle(
        'CellRight', parent=cell_style, alignment=TA_RIGHT
    )
    cell_center = ParagraphStyle(
        'CellCenter', parent=cell_style, alignment=TA_CENTER
    )
    header_style = ParagraphStyle(
        'Header', parent=styles['Normal'], fontSize=7, fontName='Helvetica-Bold',
        textColor=colors.white, alignment=TA_CENTER, leading=9
    )

    elements = []

    # Title
    from app.models import SystemSettings
    settings = SystemSettings.get_settings()
    report_date = datetime.now().strftime('%Y-%m-%d')
    elements.append(Paragraph(f'{title} — {report_date}', title_style))

    if not loans:
        elements.append(Paragraph('No active loans found.', cell_style))
        doc.build(elements)
        buf.seek(0)
        response = make_response(buf.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename={loan_frequency}_loans_{report_date}.pdf'
        return response

    # Table headers
    headers = [
        Paragraph('#', header_style),
        Paragraph('Loan Number', header_style),
        Paragraph('Customer', header_style),
        Paragraph('Phone', header_style),
        Paragraph('Loan Amount', header_style),
        Paragraph('Installment', header_style),
        Paragraph('Arrears', header_style),
        Paragraph('Monday', header_style),
        Paragraph('Tuesday', header_style),
        Paragraph('Wednesday', header_style),
        Paragraph('Thursday', header_style),
        Paragraph('Friday', header_style),
        Paragraph('Saturday', header_style),
        Paragraph('Signature', header_style),
    ]

    table_data = [headers]

    for idx, loan in enumerate(loans, 1):
        arrears = loan.get_arrears_details()
        if float(arrears['total_overdue_amount']) > 0:
            arrears_text = f"{settings.currency_symbol} {float(arrears['total_overdue_amount']):.2f}"
            n_inst = arrears['overdue_installments'] + arrears['partial_overdue_installments']
            if n_inst:
                arrears_text += f" ({n_inst})"
        else:
            arrears_text = '-'

        row = [
            Paragraph(str(idx), cell_center),
            Paragraph(esc(loan.loan_number), cell_style),
            Paragraph(f"{esc(loan.customer.full_name)}<br/><font size='5' color='grey'>{esc(loan.customer.nic_number or '')}</font>", cell_style),
            Paragraph(esc(loan.customer.phone_primary or ''), cell_style),
            Paragraph(f"{esc(settings.currency_symbol)} {float(loan.loan_amount):.2f}", cell_right),
            Paragraph(f"{esc(settings.currency_symbol)} {float(loan.installment_amount or 0):.2f}", cell_right),
            Paragraph(esc(arrears_text), cell_center),
            Paragraph('', cell_style),  # Monday
            Paragraph('', cell_style),  # Tuesday
            Paragraph('', cell_style),  # Wednesday
            Paragraph('', cell_style),  # Thursday
            Paragraph('', cell_style),  # Friday
            Paragraph('', cell_style),  # Saturday
            Paragraph('', cell_style),  # Signature
        ]
        table_data.append(row)

    # Column widths — fit landscape A4 (usable ~277mm)
    usable = page_w - 16 * mm  # total minus margins
    col_widths = [
        usable * 0.03,   # #
        usable * 0.12,   # Loan Number
        usable * 0.13,   # Customer
        usable * 0.08,   # Phone
        usable * 0.08,   # Loan Amount
        usable * 0.08,   # Installment
        usable * 0.08,   # Arrears
        usable * 0.06,   # Monday
        usable * 0.06,   # Tuesday
        usable * 0.07,   # Wednesday
        usable * 0.06,   # Thursday
        usable * 0.06,   # Friday
        usable * 0.06,   # Saturday
        usable * 0.07,   # Signature
    ]

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
        ('TOPPADDING', (0, 0), (-1, 0), 4),

        # Data rows
        ('FONTSIZE', (0, 1), (-1, -1), 6.5),
        ('TOPPADDING', (0, 1), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),

        # Grid
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),

        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),

        # Vertical alignment
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    elements.append(tbl)
    elements.append(Spacer(1, 4 * mm))
    elements.append(Paragraph(
        f"Total: {len(loans)} loan(s) — Generated {report_date}",
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7, textColor=colors.grey)
    ))

    doc.build(elements)
    buf.seek(0)

    response = make_response(buf.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={loan_frequency}_loans_{report_date}.pdf'
    return response


# Schedule Override Routes (Admin Only)
@loans_bp.route('/<int:id>/schedule', methods=['GET'])
@login_required
@admin_required
def view_schedule(id):
    """View and manage payment schedule (Admin only)"""
    loan = Loan.query.get_or_404(id)
    
    # Check branch access
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id and loan.branch_id != current_branch_id:
            flash('Access denied: Loan not found in current branch.', 'danger')
            return redirect(url_for('loans.list_loans'))
    
    # Generate payment schedule
    schedule = loan.generate_payment_schedule()
    
    return render_template('loans/schedule.html',
                         title=f'Payment Schedule: {loan.loan_number}',
                         loan=loan,
                         schedule=schedule)


@loans_bp.route('/<int:id>/schedule/override', methods=['POST'])
@login_required
@admin_required
def override_schedule(id):
    """Override/edit due date for a specific installment (Admin only)"""
    loan = Loan.query.get_or_404(id)
    
    # Check branch access
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id and loan.branch_id != current_branch_id:
            return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    try:
        data = request.get_json()
        installment_number = int(data.get('installment_number'))
        new_due_date = data.get('new_due_date')
        notes = data.get('notes', '')
        
        # Validate installment number
        schedule = loan.generate_payment_schedule()
        if installment_number < 1 or installment_number > len(schedule):
            return jsonify({'success': False, 'message': 'Invalid installment number'}), 400
        
        # Parse due date
        if new_due_date:
            new_due_date = datetime.strptime(new_due_date, '%Y-%m-%d').date()
        
        # Check if override already exists
        override = LoanScheduleOverride.query.filter_by(
            loan_id=loan.id,
            installment_number=installment_number
        ).first()
        
        if override:
            # Update existing override
            override.custom_due_date = new_due_date
            override.is_skipped = False
            override.updated_by = current_user.id
            override.updated_at = datetime.utcnow()
            if notes:
                override.notes = notes
        else:
            # Create new override
            override = LoanScheduleOverride(
                loan_id=loan.id,
                installment_number=installment_number,
                custom_due_date=new_due_date,
                is_skipped=False,
                created_by=current_user.id,
                notes=notes
            )
            db.session.add(override)
        
        # Log activity
        log = ActivityLog(
            user_id=current_user.id,
            action='override_schedule',
            entity_type='loan',
            entity_id=loan.id,
            description=f'Overrode installment {installment_number} due date to {new_due_date} for loan {loan.loan_number}',
            ip_address=request.remote_addr
        )
        db.session.add(log)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Installment {installment_number} due date updated successfully'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@loans_bp.route('/<int:id>/schedule/skip', methods=['POST'])
@login_required
@admin_required
def skip_installment(id):
    """Skip a specific installment (Admin only)"""
    loan = Loan.query.get_or_404(id)
    
    # Check branch access
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id and loan.branch_id != current_branch_id:
            return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    try:
        data = request.get_json()
        installment_number = int(data.get('installment_number'))
        notes = data.get('notes', '')
        reschedule_date_str = data.get('reschedule_date', '')
        reschedule_date = None
        if reschedule_date_str:
            from datetime import date as date_type
            reschedule_date = datetime.strptime(reschedule_date_str, '%Y-%m-%d').date()
        
        # Validate installment number
        schedule = loan.generate_payment_schedule()
        valid_installments = [inst['installment_number'] for inst in schedule]
        if installment_number not in valid_installments:
            return jsonify({'success': False, 'message': 'Invalid installment number'}), 400
        
        # Check if override already exists
        override = LoanScheduleOverride.query.filter_by(
            loan_id=loan.id,
            installment_number=installment_number
        ).first()
        
        if override:
            # Update existing override to skip
            override.is_skipped = True
            override.custom_due_date = None
            override.reschedule_date = reschedule_date
            override.updated_by = current_user.id
            override.updated_at = datetime.utcnow()
            if notes:
                override.notes = notes
        else:
            # Create new override as skipped
            override = LoanScheduleOverride(
                loan_id=loan.id,
                installment_number=installment_number,
                is_skipped=True,
                custom_due_date=None,
                reschedule_date=reschedule_date,
                created_by=current_user.id,
                notes=notes
            )
            db.session.add(override)
        
        # Log activity
        log = ActivityLog(
            user_id=current_user.id,
            action='skip_installment',
            entity_type='loan',
            entity_id=loan.id,
            description=f'Skipped installment {installment_number} for loan {loan.loan_number}' + (f', rescheduled to {reschedule_date_str}' if reschedule_date_str else ''),
            ip_address=request.remote_addr
        )
        db.session.add(log)
        
        db.session.commit()
        
        # Regenerate schedule to get computed principal/interest/status for the skipped installment
        refreshed_schedule = loan.generate_payment_schedule()
        sched_item = next((x for x in refreshed_schedule if x['installment_number'] == installment_number), None)
        
        return jsonify({
            'success': True,
            'message': f'Installment {installment_number} skipped' + (f', rescheduled to {reschedule_date_str}' if reschedule_date_str else '') + ' successfully',
            'principal': sched_item['principal'] if sched_item else None,
            'interest': sched_item['interest'] if sched_item else None,
            'status': sched_item['status'] if sched_item else 'skipped',
            'reschedule_date': reschedule_date_str or ''
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@loans_bp.route('/skip-all-daily-loans', methods=['POST'])
@login_required
@admin_required
def skip_all_daily_loans():
    """Skip one day's installment for all active daily loans (DLS & DL types)"""
    try:
        data = request.get_json()
        skip_date_str = data.get('skip_date', '')
        notes = data.get('notes', '')

        if not skip_date_str:
            return jsonify({'success': False, 'message': 'Skip date is required'}), 400

        skip_date = datetime.strptime(skip_date_str, '%Y-%m-%d').date()

        # Don't allow skipping Sundays (daily loans already skip Sundays)
        if skip_date.weekday() == 6:
            return jsonify({'success': False, 'message': 'Cannot skip a Sunday — daily loans already skip Sundays'}), 400

        # Get all active daily loans (54_daily = DLS, type4_daily = DL)
        daily_loans_query = Loan.query.filter(
            Loan.loan_type.in_(['54_daily', 'type4_daily']),
            Loan.status == 'active'
        )

        # Filter by branch if needed
        if should_filter_by_branch():
            current_branch_id = get_current_branch_id()
            if current_branch_id:
                daily_loans_query = daily_loans_query.filter_by(branch_id=current_branch_id)

        daily_loans = daily_loans_query.all()

        if not daily_loans:
            return jsonify({'success': False, 'message': 'No active daily loans found'}), 404

        skipped_count = 0
        already_skipped = 0
        not_applicable = 0
        skipped_loans = []

        for loan in daily_loans:
            schedule = loan.generate_payment_schedule()

            # Find the installment whose due_date matches the skip_date
            target_installment = None
            for inst in schedule:
                if inst['due_date'] == skip_date and not inst['is_skipped']:
                    # Only skip unpaid/pending installments
                    if inst['status'] in ('pending', 'overdue', 'partial'):
                        target_installment = inst
                        break
                elif inst['due_date'] == skip_date and inst['is_skipped']:
                    already_skipped += 1
                    break

            if not target_installment:
                not_applicable += 1
                continue

            installment_number = target_installment['installment_number']

            # Check if override already exists
            override = LoanScheduleOverride.query.filter_by(
                loan_id=loan.id,
                installment_number=installment_number
            ).first()

            if override:
                override.is_skipped = True
                override.custom_due_date = None
                override.reschedule_date = None
                override.updated_by = current_user.id
                override.updated_at = datetime.utcnow()
                if notes:
                    override.notes = notes
            else:
                override = LoanScheduleOverride(
                    loan_id=loan.id,
                    installment_number=installment_number,
                    is_skipped=True,
                    custom_due_date=None,
                    reschedule_date=None,
                    created_by=current_user.id,
                    notes=notes or f'Bulk skip for {skip_date_str}'
                )
                db.session.add(override)

            skipped_count += 1
            skipped_loans.append(loan.loan_number)

        # Log activity
        if skipped_count > 0:
            log = ActivityLog(
                user_id=current_user.id,
                action='skip_all_daily_loans',
                entity_type='loan',
                entity_id=0,
                description=f'Bulk skipped {skipped_count} daily loan installments for {skip_date_str}',
                ip_address=request.remote_addr
            )
            db.session.add(log)
            db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Successfully skipped {skipped_count} daily loan installment(s) for {skip_date_str}',
            'skipped_count': skipped_count,
            'already_skipped': already_skipped,
            'not_applicable': not_applicable,
            'skipped_loans': skipped_loans
        })

    except ValueError as e:
        return jsonify({'success': False, 'message': f'Invalid date format: {str(e)}'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@loans_bp.route('/<int:id>/schedule/reset', methods=['POST'])
@login_required
@admin_required
def reset_installment(id):
    """Reset an installment to its original due date (Admin only)"""
    loan = Loan.query.get_or_404(id)
    
    # Check branch access
    if should_filter_by_branch():
        current_branch_id = get_current_branch_id()
        if current_branch_id and loan.branch_id != current_branch_id:
            return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    try:
        data = request.get_json()
        installment_number = int(data.get('installment_number'))
        
        # Find and delete the override
        override = LoanScheduleOverride.query.filter_by(
            loan_id=loan.id,
            installment_number=installment_number
        ).first()
        
        if override:
            db.session.delete(override)
            
            # Log activity
            log = ActivityLog(
                user_id=current_user.id,
                action='reset_installment',
                entity_type='loan',
                entity_id=loan.id,
                description=f'Reset installment {installment_number} to original due date for loan {loan.loan_number}',
                ip_address=request.remote_addr
            )
            db.session.add(log)
            
            db.session.commit()
            
            # Regenerate schedule to get the original due date
            refreshed_schedule = loan.generate_payment_schedule()
            original_item = next((item for item in refreshed_schedule if item['installment_number'] == installment_number), None)
            original_due_date = original_item['due_date'].strftime('%Y-%m-%d') if original_item and original_item.get('due_date') else None
            
            return jsonify({
                'success': True,
                'message': f'Installment {installment_number} reset to original due date',
                'original_due_date': original_due_date
            })
        else:
            return jsonify({
                'success': False,
                'message': 'No override found for this installment'
            }), 404
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
