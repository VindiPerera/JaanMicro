"""Reports routes"""
from flask import render_template, request, make_response, current_app, send_from_directory, abort
from flask_login import login_required, current_user
from datetime import datetime, timedelta
from sqlalchemy import func, extract
import os
from app import db
from app.reports import reports_bp
from app.models import Customer, Loan, LoanPayment, Investment, InvestmentTransaction, Pawning, PawningPayment
from app.utils.decorators import permission_required
from app.utils.helpers import get_current_branch_id, get_branch_filter_for_query
import io
import csv

@reports_bp.route('/')
@login_required
def index():
    """Reports dashboard"""
    # Allow access if user has either view_reports or view_collection_reports permission
    if not (current_user.has_permission('view_reports') or current_user.has_permission('view_collection_reports')):
        abort(403)
    
    # Calculate quick statistics
    from decimal import Decimal
    
    # Get branch filtering info
    customer_branch_filter = get_branch_filter_for_query(Customer.branch_id)
    loan_branch_filter = get_branch_filter_for_query(Loan.branch_id)
    investment_branch_filter = get_branch_filter_for_query(Investment.branch_id)
    pawning_branch_filter = get_branch_filter_for_query(Pawning.branch_id)
    
    # Customer statistics
    customer_query = Customer.query
    if customer_branch_filter is not None:
        customer_query = customer_query.filter(customer_branch_filter)
    
    total_customers = customer_query.count()
    active_customers = customer_query.filter_by(status='active').count()
    
    # Loan statistics
    loan_query = Loan.query
    if loan_branch_filter is not None:
        loan_query = loan_query.filter(loan_branch_filter)
    
    total_loan_disbursed = db.session.query(func.sum(Loan.disbursed_amount)).filter(
        Loan.status.in_(['active', 'completed'])
    )
    if loan_branch_filter is not None:
        total_loan_disbursed = total_loan_disbursed.filter(loan_branch_filter)
    total_loan_disbursed = total_loan_disbursed.scalar() or 0
    
    active_loans = loan_query.filter_by(status='active').count()
    
    # Investment statistics
    investment_query = Investment.query
    if investment_branch_filter is not None:
        investment_query = investment_query.filter(investment_branch_filter)
    
    total_investment_amount = db.session.query(func.sum(Investment.principal_amount))
    if investment_branch_filter is not None:
        total_investment_amount = total_investment_amount.filter(investment_branch_filter)
    total_investment_amount = total_investment_amount.scalar() or 0
    
    active_investments = investment_query.filter_by(status='active').count()
    
    # Pawning statistics
    pawning_query = Pawning.query
    if pawning_branch_filter is not None:
        pawning_query = pawning_query.filter(pawning_branch_filter)
    
    active_pawnings = pawning_query.filter_by(status='active').count()
    total_pawning_amount = db.session.query(func.sum(Pawning.loan_amount)).filter_by(status='active')
    if pawning_branch_filter is not None:
        total_pawning_amount = total_pawning_amount.filter(pawning_branch_filter)
    total_pawning_amount = total_pawning_amount.scalar() or 0
    
    # Today's activity
    from datetime import date
    today = date.today()
    
    todays_loan_payments_query = db.session.query(func.sum(LoanPayment.payment_amount)).join(Loan).filter(
        LoanPayment.payment_date == today
    )
    if loan_branch_filter is not None:
        todays_loan_payments_query = todays_loan_payments_query.filter(loan_branch_filter)
    todays_loan_payments = todays_loan_payments_query.scalar() or 0
    
    todays_new_loans_query = Loan.query.filter(
        func.date(Loan.created_at) == today
    )
    if loan_branch_filter is not None:
        todays_new_loans_query = todays_new_loans_query.filter(loan_branch_filter)
    todays_new_loans = todays_new_loans_query.count()
    
    todays_new_customers_query = Customer.query.filter(
        func.date(Customer.created_at) == today
    )
    if customer_branch_filter is not None:
        todays_new_customers_query = todays_new_customers_query.filter(customer_branch_filter)
    todays_new_customers = todays_new_customers_query.count()
    
    stats = {
        'total_customers': total_customers,
        'active_customers': active_customers,
        'total_loan_disbursed': float(total_loan_disbursed),
        'active_loans': active_loans,
        'total_investment_amount': float(total_investment_amount),
        'active_investments': active_investments,
        'active_pawnings': active_pawnings,
        'total_pawning_amount': float(total_pawning_amount),
        'todays_collections': float(todays_loan_payments),
        'todays_new_loans': todays_new_loans,
        'todays_new_customers': todays_new_customers
    }
    
    return render_template('reports/index.html', title='Reports', stats=stats)

@reports_bp.route('/loans')
@login_required
@permission_required('view_reports')
def loan_report():
    """Loan reports"""
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    status = request.args.get('status', '')
    loan_purpose = request.args.get('loan_purpose', '')
    
    query = Loan.query
    
    # Apply branch filtering
    loan_branch_filter = get_branch_filter_for_query(Loan.branch_id)
    if loan_branch_filter is not None:
        query = query.filter(loan_branch_filter)
    
    if start_date:
        query = query.filter(Loan.created_at >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(Loan.created_at <= datetime.strptime(end_date, '%Y-%m-%d'))
    if status:
        query = query.filter_by(status=status)
    if loan_purpose:
        query = query.filter_by(loan_purpose=loan_purpose)
    
    loans = query.all()
    
    # Calculate payment stats for each loan
    loan_payments = {}
    for loan in loans:
        payment_stats = db.session.query(
            func.sum(LoanPayment.principal_amount),
            func.sum(LoanPayment.interest_amount)
        ).filter(LoanPayment.loan_id == loan.id)
        
        if start_date:
            payment_stats = payment_stats.filter(LoanPayment.payment_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
        if end_date:
            payment_stats = payment_stats.filter(LoanPayment.payment_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
        
        principal, interest = payment_stats.first()
        from decimal import Decimal
        principal_dec = Decimal(str(principal or 0))
        interest_dec = Decimal(str(interest or 0))
        total_dec = principal_dec + interest_dec
        
        # Calculate expected interest based on loan type
        expected_interest = loan.get_total_expected_interest()
        interest_variance = interest_dec - expected_interest
        
        loan_payments[loan.id] = {
            'principal': float(principal_dec),
            'interest': float(interest_dec),
            'total': float(total_dec),
            'expected_interest': float(expected_interest),
            'interest_variance': float(interest_variance),
            'interest_type': loan.interest_type
        }
    
    # Calculate overall statistics
    principal_collected = sum(p['principal'] for p in loan_payments.values())
    interest_collected = sum(p['interest'] for p in loan_payments.values())
    total_collected = principal_collected + interest_collected
    
    # Calculate statistics
    summary = {
        'total_loans': len(loans),
        'total_disbursed': sum(float(loan.disbursed_amount or 0) for loan in loans),
        'total_outstanding': sum(float(loan.outstanding_amount or 0) for loan in loans),
        'total_collected': total_collected,
        'principal_collected': principal_collected,
        'interest_profit': interest_collected
    }
    
    # Loan by status
    status_breakdown = db.session.query(
        Loan.status,
        func.count(Loan.id),
        func.sum(Loan.outstanding_amount)
    ).group_by(Loan.status).all()
    
    # Loan by purpose
    purpose_breakdown = db.session.query(
        Loan.loan_purpose,
        func.count(Loan.id),
        func.sum(Loan.loan_amount)
    ).group_by(Loan.loan_purpose).all()
    
    return render_template('reports/loan_report.html',
                         title='Loan Report',
                         loans=loans,
                         loan_payments=loan_payments,
                         summary=summary,
                         status_breakdown=status_breakdown,
                         purpose_breakdown=purpose_breakdown,
                         start_date=start_date,
                         end_date=end_date,
                         status=status,
                         loan_purpose=loan_purpose)

@reports_bp.route('/collections')
@login_required
@permission_required('view_collection_reports')
def collection_report():
    """Collection reports"""
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    payment_method = request.args.get('payment_method', '')
    collection_type = request.args.get('collection_type', '')
    
    # Get branch filtering info
    loan_branch_filter = get_branch_filter_for_query(Loan.branch_id)
    pawning_branch_filter = get_branch_filter_for_query(Pawning.branch_id)
    
    # Loan payments
    loan_query = LoanPayment.query.join(Loan).join(Customer)
    if loan_branch_filter is not None:
        loan_query = loan_query.filter(loan_branch_filter)
    if start_date:
        loan_query = loan_query.filter(LoanPayment.payment_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        loan_query = loan_query.filter(LoanPayment.payment_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if payment_method:
        loan_query = loan_query.filter(LoanPayment.payment_method == payment_method)
    
    loan_payments = loan_query.order_by(LoanPayment.payment_date.desc()).all() if collection_type != 'pawning' else []
    
    # Pawning payments
    pawning_query = PawningPayment.query.join(Pawning).join(Customer)
    if pawning_branch_filter is not None:
        pawning_query = pawning_query.filter(pawning_branch_filter)
    if start_date:
        pawning_query = pawning_query.filter(PawningPayment.payment_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        pawning_query = pawning_query.filter(PawningPayment.payment_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if payment_method:
        pawning_query = pawning_query.filter(PawningPayment.payment_method == payment_method)
    
    pawning_payments = pawning_query.order_by(PawningPayment.payment_date.desc()).all() if collection_type != 'loan' else []
    
    # Combine all payments
    all_payments = []
    for payment in loan_payments:
        all_payments.append({
            'payment': payment,
            'type': 'loan',
            'loan_id': payment.loan_id,
            'pawning_id': None,
            'payment_date': payment.payment_date,
            'receipt_number': payment.receipt_number,
            'reference_number': payment.loan.loan_number if payment.loan else 'N/A',
            'member_name': payment.loan.customer.full_name if payment.loan and payment.loan.customer else 'N/A',
            'amount': float(payment.payment_amount or 0),
            'principal_amount': float(payment.principal_amount or 0),
            'interest_amount': float(payment.interest_amount or 0),
            'payment_method': payment.payment_method,
            'collected_by': payment.collected_by_user if payment.collected_by_user else None
        })
    
    for payment in pawning_payments:
        all_payments.append({
            'payment': payment,
            'type': 'pawning',
            'loan_id': None,
            'pawning_id': payment.pawning_id,
            'payment_date': payment.payment_date,
            'receipt_number': payment.receipt_number,
            'reference_number': payment.pawning.ticket_number if payment.pawning else 'N/A',
            'member_name': payment.pawning.customer.full_name if payment.pawning and payment.pawning.customer else 'N/A',
            'amount': float(payment.payment_amount or 0),
            'principal_amount': float(payment.principal_amount or 0),
            'interest_amount': float(payment.interest_amount or 0),
            'payment_method': payment.payment_method,
            'collected_by': payment.collected_by_user if payment.collected_by_user else None
        })
    
    # Sort all payments by date descending
    all_payments.sort(key=lambda x: x['payment_date'], reverse=True)
    
    # Calculate summary
    total_amount = sum(p['amount'] for p in all_payments)
    total_principal = sum(p['principal_amount'] for p in all_payments)
    total_interest = sum(p['interest_amount'] for p in all_payments)
    
    summary = {
        'total_count': len(all_payments),
        'total_amount': total_amount,
        'total_principal': total_principal,
        'total_interest': total_interest
    }
    
    # Collections by payment method
    collections_by_method = {}
    for payment in all_payments:
        method = payment['payment_method'] or 'Not Specified'
        if method not in collections_by_method:
            collections_by_method[method] = {'count': 0, 'total': 0}
        collections_by_method[method]['count'] += 1
        collections_by_method[method]['total'] += payment['amount']
    
    collections_by_method_list = [
        {'payment_method': k, 'count': v['count'], 'total': v['total']}
        for k, v in collections_by_method.items()
    ]
    
    # Collections by user
    collections_by_user = {}
    for payment in all_payments:
        user = payment['collected_by']
        user_name = user.full_name if user else 'Not Specified'
        if user_name not in collections_by_user:
            collections_by_user[user_name] = {'count': 0, 'total': 0}
        collections_by_user[user_name]['count'] += 1
        collections_by_user[user_name]['total'] += payment['amount']
    
    collections_by_user_list = [
        {'user_name': k, 'count': v['count'], 'total': v['total']}
        for k, v in collections_by_user.items()
    ]
    
    return render_template('reports/collection_report.html',
                         title='Collection Report',
                         payments=all_payments,
                         summary=summary,
                         collections_by_method=collections_by_method_list,
                         collections_by_user=collections_by_user_list,
                         start_date=start_date,
                         end_date=end_date,
                         payment_method=payment_method,
                         collection_type=collection_type)

@reports_bp.route('/customers')
@login_required
@permission_required('view_reports')
def customer_report():
    """Customer reports"""
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    status = request.args.get('status', '')
    kyc_status = request.args.get('kyc_status', '')
    district = request.args.get('district', '')
    
    query = Customer.query
    
    # Apply branch filtering
    customer_branch_filter = get_branch_filter_for_query(Customer.branch_id)
    if customer_branch_filter is not None:
        query = query.filter(customer_branch_filter)
    
    # Date filtering
    if start_date:
        query = query.filter(Customer.created_at >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        # Add one day to include the end date
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(Customer.created_at < end_datetime)
    
    # Status filtering
    if status:
        query = query.filter_by(status=status)
    if kyc_status == 'verified':
        query = query.filter_by(kyc_verified=True)
    elif kyc_status == 'pending':
        query = query.filter_by(kyc_verified=False)
    
    # District filtering
    if district:
        query = query.filter_by(district=district)
    
    customers = query.all()
    
    # Add counts for each customer (also filtered by branch)
    for customer in customers:
        loan_count_query = Loan.query.filter_by(customer_id=customer.id, status='active')
        investment_count_query = Investment.query.filter_by(customer_id=customer.id, status='active')
        pawning_count_query = Pawning.query.filter_by(customer_id=customer.id, status='active')
        
        if customer_branch_filter is not None:
            loan_count_query = loan_count_query.filter(customer_branch_filter)
            investment_count_query = investment_count_query.filter(customer_branch_filter)
            pawning_count_query = pawning_count_query.filter(customer_branch_filter)
        
        customer.active_loans_count = loan_count_query.count()
        customer.active_investments_count = investment_count_query.count()
        customer.active_pawnings_count = pawning_count_query.count()
    
    # Statistics
    summary = {
        'total_customers': query.count(),
        'active_customers': query.filter_by(status='active').count(),
        'kyc_verified': query.filter_by(kyc_verified=True).count(),
        'kyc_pending': query.filter_by(kyc_verified=False).count(),
    }
    
    # Get counts for customers with active products (filtered by branch)
    loan_customer_query = db.session.query(func.count(func.distinct(Loan.customer_id)))
    investment_customer_query = db.session.query(func.count(func.distinct(Investment.customer_id)))
    pawning_customer_query = db.session.query(func.count(func.distinct(Pawning.customer_id)))
    
    if customer_branch_filter is not None:
        loan_customer_query = loan_customer_query.filter(customer_branch_filter)
        investment_customer_query = investment_customer_query.filter(customer_branch_filter)
        pawning_customer_query = pawning_customer_query.filter(customer_branch_filter)
    
    summary.update({
        'customers_with_loans': loan_customer_query.filter(Loan.status == 'active').scalar() or 0,
        'customers_with_investments': investment_customer_query.filter(Investment.status == 'active').scalar() or 0,
        'customers_with_pawnings': pawning_customer_query.filter(Pawning.status == 'active').scalar() or 0
    })
    
    # Geographic distribution
    district_query = db.session.query(
        Customer.district,
        func.count(Customer.id).label('count')
    )
    if customer_branch_filter is not None:
        district_query = district_query.filter(customer_branch_filter)
    district_breakdown = district_query.group_by(Customer.district).order_by(func.count(Customer.id).desc()).limit(10).all()
    
    # Occupation distribution
    occupation_query = db.session.query(
        Customer.occupation,
        func.count(Customer.id).label('count')
    )
    if customer_branch_filter is not None:
        occupation_query = occupation_query.filter(customer_branch_filter)
    occupation_breakdown = occupation_query.group_by(Customer.occupation).order_by(func.count(Customer.id).desc()).limit(10).all()
    
    # Get all available districts for filter dropdown
    districts_query = db.session.query(Customer.district).distinct().filter(Customer.district != None)
    if customer_branch_filter is not None:
        districts_query = districts_query.filter(customer_branch_filter)
    available_districts = districts_query.order_by(Customer.district).all()
    available_districts = [d[0] for d in available_districts if d[0]]  # Convert to list and filter out None values
    
    return render_template('reports/customer_report.html',
                         title='Member Report',
                         customers=customers,
                         summary=summary,
                         district_breakdown=district_breakdown,
                         occupation_breakdown=occupation_breakdown,
                         available_districts=available_districts,
                         start_date=start_date,
                         end_date=end_date,
                         status=status,
                         kyc_status=kyc_status,
                         district=district)

@reports_bp.route('/investments')
@login_required
@permission_required('view_reports')
def investment_report():
    """Borrower reports"""
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    investment_type = request.args.get('investment_type', '')
    
    query = Investment.query
    
    # Apply branch filtering
    investment_branch_filter = get_branch_filter_for_query(Investment.branch_id)
    if investment_branch_filter is not None:
        query = query.filter(investment_branch_filter)
    
    if start_date:
        query = query.filter(Investment.created_at >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(Investment.created_at <= datetime.strptime(end_date, '%Y-%m-%d'))
    if investment_type:
        query = query.filter_by(investment_type=investment_type)
    
    investments = query.all()
    
    # Calculate interest paid/accrued (this is expense for the company)
    total_interest_expense = sum(float(inv.current_amount - inv.principal_amount) for inv in investments if inv.current_amount and inv.current_amount > inv.principal_amount)
    total_current = sum(float(inv.current_amount) for inv in investments if inv.current_amount)
    total_maturity = sum(float(inv.maturity_amount) for inv in investments if inv.maturity_amount)
    
    # Statistics
    summary = {
        'total_investments': len(investments),
        'total_principal': sum(float(inv.principal_amount) for inv in investments),
        'current_amount': total_current,
        'total_maturity': total_maturity,
        'total_interest_paid': total_interest_expense,
        'interest_expense': total_interest_expense  # This is a cost/expense
    }
    
    # Type breakdown
    type_breakdown_query = db.session.query(
        Investment.investment_type,
        func.count(Investment.id),
        func.sum(Investment.current_amount)
    )
    if investment_branch_filter is not None:
        type_breakdown_query = type_breakdown_query.filter(investment_branch_filter)
    type_breakdown = type_breakdown_query.group_by(Investment.investment_type).all()
    
    # Get investments by type for display
    investments_by_type = []
    for inv_type, count, total in type_breakdown:
        investments_by_type.append({
            'investment_type': inv_type,
            'count': count,
            'total': float(total or 0)
        })
    
    # Get maturing investments (within next 30 days)
    from datetime import date, timedelta
    thirty_days = date.today() + timedelta(days=30)
    maturing_query = Investment.query.filter(
        Investment.maturity_date <= thirty_days,
        Investment.status == 'active'
    )
    if investment_branch_filter is not None:
        maturing_query = maturing_query.filter(investment_branch_filter)
    maturing_soon = maturing_query.all()
    
    return render_template('reports/investment_report.html',
                         title='Borrower Report',
                         investments=investments,
                         summary=summary,
                         type_breakdown=type_breakdown,
                         investments_by_type=investments_by_type,
                         maturing_soon=maturing_soon,
                         start_date=start_date,
                         end_date=end_date,
                         investment_type=investment_type)

@reports_bp.route('/pawnings')
@login_required
@permission_required('view_reports')
def pawning_report():
    """Pawning reports"""
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    status = request.args.get('status', '')
    item_type = request.args.get('item_type', '')
    
    query = Pawning.query
    
    # Apply branch filtering
    pawning_branch_filter = get_branch_filter_for_query(Pawning.branch_id)
    if pawning_branch_filter is not None:
        query = query.filter(pawning_branch_filter)
    
    if start_date:
        query = query.filter(Pawning.pawning_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Pawning.pawning_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if status:
        query = query.filter_by(status=status)
    if item_type:
        query = query.filter_by(item_type=item_type)
    
    pawnings = query.all()
    
    # Calculate total interest collected from pawning payments
    total_interest_collected = db.session.query(func.sum(PawningPayment.interest_amount)).join(
        Pawning, PawningPayment.pawning_id == Pawning.id
    )
    if pawning_branch_filter is not None:
        total_interest_collected = total_interest_collected.filter(pawning_branch_filter)
    if start_date:
        total_interest_collected = total_interest_collected.filter(PawningPayment.payment_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        total_interest_collected = total_interest_collected.filter(PawningPayment.payment_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if status:
        total_interest_collected = total_interest_collected.filter(Pawning.status == status)
    
    interest_profit = float(total_interest_collected.scalar() or 0)
    
    # Statistics
    from decimal import Decimal
    total_loan = sum(Decimal(str(pawn.loan_amount)) for pawn in pawnings)
    total_outstanding = sum(Decimal(str(pawn.outstanding_principal or 0)) for pawn in pawnings if pawn.status == 'active')
    total_collected = sum(Decimal(str(pawn.principal_paid or 0)) + Decimal(str(pawn.total_interest_paid or 0)) for pawn in pawnings)
    
    summary = {
        'total_pawnings': len(pawnings),
        'total_loan_amount': float(total_loan),
        'total_outstanding': float(total_outstanding),
        'total_collected': float(total_collected),
        'interest_profit': interest_profit
    }
    
    # Status breakdown
    status_breakdown_query = db.session.query(
        Pawning.status,
        func.count(Pawning.id),
        func.sum(Pawning.outstanding_principal)
    )
    if pawning_branch_filter is not None:
        status_breakdown_query = status_breakdown_query.filter(pawning_branch_filter)
    status_breakdown = status_breakdown_query.group_by(Pawning.status).all()
    
    # Get overdue pawnings
    from datetime import date
    today = date.today()
    overdue_query = Pawning.query.filter(
        Pawning.maturity_date < today,
        Pawning.status == 'active'
    )
    if pawning_branch_filter is not None:
        overdue_query = overdue_query.filter(pawning_branch_filter)
    overdue_pawnings = overdue_query.all()
    
    return render_template('reports/pawning_report.html',
                         title='Pawning Report',
                         pawnings=pawnings,
                         summary=summary,
                         status_breakdown=status_breakdown,
                         overdue_pawnings=overdue_pawnings,
                         today=today,
                         start_date=start_date,
                         end_date=end_date,
                         status=status,
                         item_type=item_type)

@reports_bp.route('/arrears')
@login_required
@permission_required('view_reports')
def arrears_report():
    """Arrears report - overdue amounts that customers need to pay
    Includes: loans past maturity AND loans with overdue/partial installments before maturity
    """
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    status = request.args.get('status', 'active')  # Default to active
    product_type = request.args.get('product_type', '')  # loan, pawning, or all
    
    # Get branch filtering info
    current_branch_id = get_current_branch_id()
    
    from datetime import date
    from decimal import Decimal, ROUND_HALF_UP
    today = date.today()
    
    arrears_data = []
    
    # Loan arrears - only active loans past maturity date
    if product_type in ['', 'loan']:
        loan_query = Loan.query.filter_by(status='active')
        
        loan_branch_filter = get_branch_filter_for_query(Loan.branch_id)
        if loan_branch_filter is not None:
            loan_query = loan_query.filter(loan_branch_filter)
        
        if start_date:
            loan_query = loan_query.filter(Loan.disbursement_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
        if end_date:
            loan_query = loan_query.filter(Loan.disbursement_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
        
        loans = loan_query.all()
        
        for loan in loans:
            from decimal import Decimal, ROUND_HALF_UP
            
            # Get arrears details using the schedule-based method
            arrears_details = loan.get_arrears_details()
            total_overdue_amount = arrears_details['total_overdue_amount']
            
            # Skip loans with no overdue amounts (neither past maturity nor installment overdue)
            if total_overdue_amount <= Decimal('0'):
                continue
            
            # Calculate outstanding amounts
            disbursed = Decimal(str(loan.disbursed_amount or loan.loan_amount))
            outstanding_principal = Decimal(str(loan.outstanding_amount or 0))
            
            # For display: principal portion of overdue vs interest portion
            # Use the installment amount split to estimate
            arrears_principal = outstanding_principal
            arrears_interest = Decimal('0')
            
            # Total arrears
            penalty = Decimal(str(loan.penalty_amount or 0))
            advance_balance = Decimal(str(loan.advance_balance or 0))
            
            # Determine arrears type: past maturity vs installment overdue
            is_past_maturity = bool(loan.maturity_date and loan.maturity_date < today)
            overdue_days = arrears_details['days_overdue']
            oldest_overdue_date = arrears_details['oldest_overdue_date']
            
            # The overdue amount from schedule (includes partial remainders)
            installment_overdue_amount = float(total_overdue_amount)
            num_arrears = arrears_details['overdue_installments'] + arrears_details['partial_overdue_installments']
            
            arrears_data.append({
                'product_type': 'Loan',
                'reference_number': loan.loan_number,
                'loan_id': loan.id,
                'customer_name': loan.customer.full_name,
                'customer_id': loan.customer.customer_id,
                'customer_phone': loan.customer.phone_primary,
                'customer_nic': loan.customer.nic_number,
                'customer_address': f"{loan.customer.address_line1}, {loan.customer.city}, {loan.customer.district}",
                'disbursement_date': loan.disbursement_date,
                'maturity_date': loan.maturity_date,
                'original_amount': float(disbursed),
                'principal_outstanding': float(arrears_principal),
                'interest_outstanding': float(arrears_interest),
                'penalty': float(penalty),
                'installment_overdue': installment_overdue_amount,
                'overdue_installments': arrears_details['overdue_installments'],
                'partial_overdue_amount': float(arrears_details['partial_overdue_amount']),
                'partial_overdue_installments': arrears_details['partial_overdue_installments'],
                'advance_balance': float(advance_balance),
                'outstanding': float(outstanding_principal),
                'overdue_amount': installment_overdue_amount,
                'total_arrears': installment_overdue_amount,
                'num_arrears': num_arrears,
                'is_overdue': True,
                'is_past_maturity': is_past_maturity,
                'days_overdue': overdue_days,
                'oldest_overdue_date': oldest_overdue_date,
                'loan_type': loan.loan_type,
                'interest_type': loan.interest_type
            })
    
    # Pawning arrears - only active pawnings past maturity date
    if product_type in ['', 'pawning']:
        pawning_query = Pawning.query.filter_by(status='active')
        
        pawning_branch_filter = get_branch_filter_for_query(Pawning.branch_id)
        if pawning_branch_filter is not None:
            pawning_query = pawning_query.filter(pawning_branch_filter)
        
        if start_date:
            pawning_query = pawning_query.filter(Pawning.pawning_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
        if end_date:
            pawning_query = pawning_query.filter(Pawning.pawning_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
        
        pawnings = pawning_query.all()
        
        # Filter pawnings that are actually overdue (past maturity date)
        overdue_pawnings = []
        for pawning in pawnings:
            maturity_date = pawning.extended_date or pawning.maturity_date
            if maturity_date and maturity_date < today:
                overdue_pawnings.append(pawning)
        
        pawnings = overdue_pawnings
        
        for pawning in pawnings:
            loan_amount = Decimal(str(pawning.loan_amount))
            principal_paid = Decimal(str(pawning.principal_paid or 0))
            outstanding_principal = loan_amount - principal_paid
            interest_due = Decimal(str(pawning.interest_due or 0))
            penalty = Decimal(str(pawning.total_penalty or 0))
            total_arrears = outstanding_principal + interest_due + penalty
            
            overdue_days = 0
            maturity_date = pawning.extended_date or pawning.maturity_date
            if maturity_date:
                days_diff = (today - maturity_date).days
                overdue_days = max(0, days_diff)
            
            arrears_data.append({
                'product_type': 'Pawning',
                'reference_number': pawning.pawning_number,
                'loan_id': None,
                'customer_name': pawning.customer.full_name,
                'customer_id': pawning.customer.customer_id,
                'customer_phone': pawning.customer.phone_primary,
                'customer_nic': pawning.customer.nic_number,
                'customer_address': f"{pawning.customer.address_line1}, {pawning.customer.city}, {pawning.customer.district}",
                'disbursement_date': pawning.pawning_date,
                'maturity_date': maturity_date,
                'original_amount': float(loan_amount),
                'principal_outstanding': float(outstanding_principal),
                'interest_outstanding': float(interest_due),
                'penalty': float(penalty),
                'installment_overdue': float(total_arrears),
                'overdue_installments': 0,
                'partial_overdue_amount': 0,
                'partial_overdue_installments': 0,
                'advance_balance': 0,
                'outstanding': float(outstanding_principal + interest_due),
                'overdue_amount': float(total_arrears),
                'total_arrears': float(total_arrears),
                'num_arrears': 1,
                'is_overdue': True,
                'is_past_maturity': True,
                'days_overdue': overdue_days,
                'oldest_overdue_date': maturity_date,
                'loan_type': pawning.item_type,
                'interest_type': 'monthly'
            })
    
    # Sort by number of arrears descending, then total arrears
    arrears_data.sort(key=lambda x: (x['num_arrears'], x['total_arrears']), reverse=True)
    
    # Calculate summary statistics
    total_arrears = sum(Decimal(str(item['total_arrears'])) for item in arrears_data)
    total_principal = sum(Decimal(str(item['principal_outstanding'])) for item in arrears_data)
    total_interest = sum(Decimal(str(item['interest_outstanding'])) for item in arrears_data)
    total_penalty = sum(Decimal(str(item['penalty'])) for item in arrears_data)
    total_installment_overdue = sum(Decimal(str(item['installment_overdue'])) for item in arrears_data)
    total_partial = sum(Decimal(str(item['partial_overdue_amount'])) for item in arrears_data)
    total_num_arrears = sum(item['num_arrears'] for item in arrears_data)
    
    past_maturity_items = [item for item in arrears_data if item.get('is_past_maturity')]
    installment_only_items = [item for item in arrears_data if not item.get('is_past_maturity')]
    
    summary = {
        'total_accounts': len(arrears_data),
        'total_arrears': float(total_arrears),
        'total_principal': float(total_principal),
        'total_interest': float(total_interest),
        'total_penalty': float(total_penalty),
        'total_installment_overdue': float(total_installment_overdue),
        'total_partial_overdue': float(total_partial),
        'total_num_arrears': total_num_arrears,
        'overdue_accounts': len(arrears_data),
        'overdue_amount': float(total_arrears),
        'past_maturity_accounts': len(past_maturity_items),
        'installment_overdue_accounts': len(installment_only_items)
    }
    
    # Breakdown by product type
    loan_arrears = [item for item in arrears_data if item['product_type'] == 'Loan']
    pawning_arrears = [item for item in arrears_data if item['product_type'] == 'Pawning']
    
    product_breakdown = [
        {
            'product_type': 'Loan',
            'count': len(loan_arrears),
            'total_arrears': sum(item['total_arrears'] for item in loan_arrears)
        },
        {
            'product_type': 'Pawning',
            'count': len(pawning_arrears),
            'total_arrears': sum(item['total_arrears'] for item in pawning_arrears)
        }
    ]
    
    return render_template('reports/arrears_report.html',
                         title='Arrears Report',
                         arrears_data=arrears_data,
                         summary=summary,
                         product_breakdown=product_breakdown,
                         start_date=start_date,
                         end_date=end_date,
                         status=status,
                         product_type=product_type)


@reports_bp.route('/documentation-charges')
@login_required
@permission_required('view_reports')
def documentation_charges_report():
    """Documentation charges report"""
    from decimal import Decimal
    
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    loan_type = request.args.get('loan_type', '')
    
    query = Loan.query.filter(Loan.status.in_(['active', 'completed', 'pending', 'initiated']))
    
    # Apply branch filtering
    loan_branch_filter = get_branch_filter_for_query(Loan.branch_id)
    if loan_branch_filter is not None:
        query = query.filter(loan_branch_filter)
    
    if start_date:
        query = query.filter(Loan.created_at >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(Loan.created_at <= datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1))
    if loan_type:
        query = query.filter(Loan.loan_type == loan_type)
    
    loans = query.order_by(Loan.created_at.desc()).all()
    
    # Calculate summary
    total_doc_fees = sum(float(loan.documentation_fee or 0) for loan in loans)
    total_loan_amount = sum(float(loan.loan_amount or 0) for loan in loans)
    total_loans = len(loans)
    
    summary = {
        'total_loans': total_loans,
        'total_doc_fees': total_doc_fees,
        'total_loan_amount': total_loan_amount,
    }
    
    return render_template('reports/documentation_charges_report.html',
                         title='Documentation Charges Report',
                         loans=loans,
                         summary=summary,
                         start_date=start_date,
                         end_date=end_date,
                         loan_type=loan_type)


@reports_bp.route('/export/documentation-charges')
@login_required
@permission_required('view_reports')
def export_documentation_charges():
    """Export documentation charges report to Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    loan_type = request.args.get('loan_type', '')
    
    query = Loan.query.filter(Loan.status.in_(['active', 'completed', 'pending', 'initiated']))
    
    loan_branch_filter = get_branch_filter_for_query(Loan.branch_id)
    if loan_branch_filter is not None:
        query = query.filter(loan_branch_filter)
    
    if start_date:
        query = query.filter(Loan.created_at >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(Loan.created_at <= datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1))
    if loan_type:
        query = query.filter(Loan.loan_type == loan_type)
    
    loans = query.order_by(Loan.created_at.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Documentation Charges'
    
    headers = [
        'Loan Number', 'Customer Name', 'NIC Number', 'Loan Type',
        'Loan Amount', 'Documentation Fee', 'Disbursed Amount',
        'Status', 'Created Date'
    ]
    
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    total_fees = 0
    total_amount = 0
    
    for loan in loans:
        doc_fee = float(loan.documentation_fee or 0)
        loan_amt = float(loan.loan_amount or 0)
        total_fees += doc_fee
        total_amount += loan_amt
        
        ws.append([
            loan.loan_number,
            loan.customer.full_name if loan.customer else 'N/A',
            loan.customer.nic_number if loan.customer else 'N/A',
            loan.loan_type.replace('_', ' ').title() if loan.loan_type else 'N/A',
            loan_amt,
            doc_fee,
            float(loan.disbursed_amount or 0),
            loan.status.title() if loan.status else 'N/A',
            loan.created_at.strftime('%Y-%m-%d') if loan.created_at else 'N/A'
        ])
    
    # Add totals row
    ws.append([])
    total_row = ['', '', '', 'TOTAL', total_amount, total_fees, '', '', '']
    ws.append(total_row)
    bold_font = Font(bold=True)
    for cell in ws[ws.max_row]:
        cell.font = bold_font
    
    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=documentation_charges_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    return response


@reports_bp.route('/export/loans')
@login_required
@permission_required('view_reports')
def export_loans():
    """Export all loans to Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    query = Loan.query

    # Apply branch filtering
    loan_branch_filter = get_branch_filter_for_query(Loan.branch_id)
    if loan_branch_filter is not None:
        query = query.filter(loan_branch_filter)

    loans = query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Loan Report'

    headers = [
        'Loan Number', 'Customer', 'Loan Purpose', 'Calculation Type',
        'Disbursement Date', 'Loan Amount', 'Interest Rate', 'Installment Amount',
        'Duration', 'Outstanding Amount', 'Status', 'Referred By', 'Created Date'
    ]

    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for loan in loans:
        if loan.duration_weeks:
            duration = f"{loan.duration_weeks} weeks"
        elif loan.duration_months:
            duration = f"{loan.duration_months} months"
        else:
            duration = 'N/A'

        referred_by_name = loan.referrer.full_name if loan.referrer else 'N/A'

        ws.append([
            loan.loan_number,
            loan.customer.full_name if loan.customer else 'N/A',
            loan.loan_purpose or 'N/A',
            loan.loan_type or 'N/A',
            loan.disbursement_date.strftime('%Y-%m-%d') if loan.disbursement_date else 'N/A',
            float(loan.loan_amount) if loan.loan_amount else 0,
            float(loan.interest_rate) if loan.interest_rate else 0,
            float(loan.installment_amount) if loan.installment_amount else 0,
            duration,
            float(loan.outstanding_amount) if loan.outstanding_amount else 0,
            loan.status or 'N/A',
            referred_by_name,
            loan.created_at.strftime('%Y-%m-%d') if loan.created_at else 'N/A'
        ])

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=loans_{datetime.now().strftime("%Y%m%d")}.xlsx'

    return response

@reports_bp.route('/export/arrears')
@login_required
@permission_required('view_reports')
def export_arrears():
    """Export arrears report to CSV with full detail"""
    from datetime import date
    from decimal import Decimal, ROUND_HALF_UP
    today = date.today()
    
    product_type = request.args.get('product_type', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    arrears_data = []
    
    # Loan arrears
    if product_type in ['', 'loan']:
        loan_query = Loan.query.filter_by(status='active')
        loan_branch_filter = get_branch_filter_for_query(Loan.branch_id)
        if loan_branch_filter is not None:
            loan_query = loan_query.filter(loan_branch_filter)
        if start_date:
            loan_query = loan_query.filter(Loan.disbursement_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
        if end_date:
            loan_query = loan_query.filter(Loan.disbursement_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
        
        for loan in loan_query.all():
            details = loan.get_arrears_details()
            total_overdue = details['total_overdue_amount']
            if total_overdue <= Decimal('0'):
                continue
            
            disbursed = Decimal(str(loan.disbursed_amount or loan.loan_amount))
            is_past_maturity = loan.maturity_date and loan.maturity_date < today
            
            arrears_data.append({
                'type': 'Loan',
                'reference': loan.loan_number,
                'customer': loan.customer.full_name,
                'customer_id': loan.customer.customer_id,
                'phone': loan.customer.phone_primary,
                'nic': loan.customer.nic_number,
                'address': f"{loan.customer.address_line1}, {loan.customer.city}, {loan.customer.district}",
                'loan_type': loan.loan_type,
                'disbursement_date': loan.disbursement_date.strftime('%Y-%m-%d') if loan.disbursement_date else 'N/A',
                'maturity_date': loan.maturity_date.strftime('%Y-%m-%d') if loan.maturity_date else 'N/A',
                'original_amount': float(disbursed),
                'outstanding': float(loan.outstanding_amount or 0),
                'overdue_installments': details['overdue_installments'],
                'partial_installments': details['partial_overdue_installments'],
                'overdue_amount': float(total_overdue),
                'partial_overdue': float(details['partial_overdue_amount']),
                'advance_balance': float(loan.advance_balance or 0),
                'days_overdue': details['days_overdue'],
                'status': 'Past Maturity' if is_past_maturity else 'Installment Overdue',
            })
    
    # Pawning arrears
    if product_type in ['', 'pawning']:
        pawning_query = Pawning.query.filter_by(status='active')
        pawning_branch_filter = get_branch_filter_for_query(Pawning.branch_id)
        if pawning_branch_filter is not None:
            pawning_query = pawning_query.filter(pawning_branch_filter)
        
        for pawning in pawning_query.all():
            maturity_date = pawning.extended_date or pawning.maturity_date
            if not (maturity_date and maturity_date < today):
                continue
            
            loan_amount = Decimal(str(pawning.loan_amount))
            principal_paid = Decimal(str(pawning.principal_paid or 0))
            outstanding_principal = loan_amount - principal_paid
            interest_due = Decimal(str(pawning.interest_due or 0))
            penalty = Decimal(str(pawning.total_penalty or 0))
            total_arrears = outstanding_principal + interest_due + penalty
            overdue_days = max(0, (today - maturity_date).days) if maturity_date else 0
            
            arrears_data.append({
                'type': 'Pawning',
                'reference': pawning.pawning_number,
                'customer': pawning.customer.full_name,
                'customer_id': pawning.customer.customer_id,
                'phone': pawning.customer.phone_primary,
                'nic': pawning.customer.nic_number,
                'address': f"{pawning.customer.address_line1}, {pawning.customer.city}, {pawning.customer.district}",
                'loan_type': pawning.item_type,
                'disbursement_date': pawning.pawning_date.strftime('%Y-%m-%d') if pawning.pawning_date else 'N/A',
                'maturity_date': maturity_date.strftime('%Y-%m-%d') if maturity_date else 'N/A',
                'original_amount': float(loan_amount),
                'outstanding': float(outstanding_principal + interest_due),
                'overdue_installments': 0,
                'partial_installments': 0,
                'overdue_amount': float(total_arrears),
                'partial_overdue': 0,
                'advance_balance': 0,
                'days_overdue': overdue_days,
                'status': 'Past Maturity',
            })
    
    arrears_data.sort(key=lambda x: x['overdue_amount'], reverse=True)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow([
        'Type', 'Reference #', 'Customer', 'Member ID', 'Phone', 'NIC', 'Address',
        'Loan Type', 'Disbursement Date', 'Settlement Date', 'Original Amount',
        'Outstanding', 'Overdue Installments', 'Partial Installments',
        'Overdue Amount', 'Partial Overdue', 'Advance Balance', 'Days Overdue', 'Status'
    ])
    
    for item in arrears_data:
        writer.writerow([
            item['type'], item['reference'], item['customer'], item['customer_id'],
            item['phone'], item['nic'], item['address'], item['loan_type'],
            item['disbursement_date'], item['maturity_date'], item['original_amount'],
            item['outstanding'], item['overdue_installments'], item['partial_installments'],
            item['overdue_amount'], item['partial_overdue'], item['advance_balance'],
            item['days_overdue'], item['status']
        ])
    
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=arrears_report_{datetime.now().strftime("%Y%m%d")}.csv'
    
    return response

@reports_bp.route('/customer/<int:customer_id>/kyc/<document_type>')
@login_required
@permission_required('view_reports')
def view_kyc_document(customer_id, document_type):
    """View customer KYC documents"""
    customer = Customer.query.get_or_404(customer_id)
    
    # Define allowed document types and their corresponding fields
    document_fields = {
        'nic_front': customer.nic_front_image,
        'nic_back': customer.nic_back_image,
        'photo': customer.photo,
        'proof_of_address': customer.proof_of_address
    }
    
    if document_type not in document_fields:
        abort(404)
    
    document_path = document_fields[document_type]
    if not document_path:
        abort(404, description="Document not found")
    
    # Strip leading 'uploads/' if present (handles old DB records)
    if document_path.startswith('uploads/'):
        document_path = document_path[len('uploads/'):]
    
    # Construct full path
    upload_folder = current_app.config.get('UPLOAD_FOLDER', 'app/static/uploads')
    full_path = os.path.join(upload_folder, document_path)
    
    if not os.path.exists(full_path):
        abort(404, description="Document file not found")
    
    # Get directory and filename
    directory = os.path.dirname(full_path)
    filename = os.path.basename(full_path)
    
    return send_from_directory(directory, filename)


@reports_bp.route('/daily-installments')
@login_required
@permission_required('view_reports')
def daily_installments_report():
    """Daily installments report — shows every loan installment due within a date range."""
    from datetime import date
    from decimal import Decimal

    today = date.today()
    start_date_str = request.args.get('start_date', today.strftime('%Y-%m-%d'))
    end_date_str   = request.args.get('end_date',   today.strftime('%Y-%m-%d'))
    loan_type_filter = request.args.get('loan_type', '')

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date   = datetime.strptime(end_date_str,   '%Y-%m-%d').date()
    except ValueError:
        start_date = end_date = today

    # Active loans (status active)
    loan_query = Loan.query.filter_by(status='active')
    loan_branch_filter = get_branch_filter_for_query(Loan.branch_id)
    if loan_branch_filter is not None:
        loan_query = loan_query.filter(loan_branch_filter)
    if loan_type_filter:
        loan_query = loan_query.filter_by(loan_type=loan_type_filter)

    loans = loan_query.all()

    # Build guarantor lookup: {customer_id: [loan_numbers they guarantee]}
    # We also pre-fetch guarantor Customer objects keyed by id
    from app.models import Customer as CustomerModel
    all_customer_ids = set()
    for loan in loans:
        if loan.guarantor_ids:
            for gid in loan.guarantor_ids.split(','):
                gid = gid.strip()
                if gid:
                    all_customer_ids.add(int(gid))

    guarantor_map = {}
    if all_customer_ids:
        gs = CustomerModel.query.filter(CustomerModel.id.in_(all_customer_ids)).all()
        guarantor_map = {g.id: g for g in gs}

    rows = []
    summary_total_amount   = Decimal('0')
    summary_total_paid     = Decimal('0')
    summary_total_remaining = Decimal('0')
    overdue_count  = 0
    pending_count  = 0
    paid_count     = 0
    partial_count  = 0

    for loan in loans:
        schedule = loan.generate_payment_schedule()
        payments = loan.payments.order_by(LoanPayment.payment_date.desc()).all()

        # Build guarantors list for this loan
        guarantors = []
        if loan.guarantor_ids:
            for gid in loan.guarantor_ids.split(','):
                gid = gid.strip()
                if gid and int(gid) in guarantor_map:
                    guarantors.append(guarantor_map[int(gid)])

        for inst in schedule:
            due = inst['due_date']
            if not (start_date <= due <= end_date):
                continue
            if inst.get('is_skipped'):
                continue

            amount    = Decimal(str(inst['amount']))
            paid_amt  = Decimal(str(inst['paid_amount']))
            remaining = Decimal(str(inst['remaining_amount']))
            status    = inst['status']

            summary_total_amount    += amount
            summary_total_paid      += paid_amt
            summary_total_remaining += remaining

            if status == 'overdue':    overdue_count  += 1
            elif status == 'pending':  pending_count  += 1
            elif status == 'paid':     paid_count     += 1
            elif status == 'partial':  partial_count  += 1

            rows.append({
                'loan': loan,
                'installment_number': inst['installment_number'],
                'due_date': due,
                'amount': float(amount),
                'principal': inst['principal'],
                'interest': inst['interest'],
                'paid_amount': float(paid_amt),
                'remaining_amount': float(remaining),
                'status': status,
                'guarantors': guarantors,
                'payments': payments,  # all payments for this loan (history)
            })

    # Sort by due_date, then loan_number
    rows.sort(key=lambda r: (r['due_date'], r['loan'].loan_number))

    summary = {
        'total_installments': len(rows),
        'total_amount':    float(summary_total_amount),
        'total_paid':      float(summary_total_paid),
        'total_remaining': float(summary_total_remaining),
        'overdue_count':  overdue_count,
        'pending_count':  pending_count,
        'paid_count':     paid_count,
        'partial_count':  partial_count,
    }

    loan_types = db.session.query(Loan.loan_type).distinct().order_by(Loan.loan_type).all()
    loan_types = [lt[0] for lt in loan_types if lt[0]]

    return render_template(
        'reports/daily_installments.html',
        title='Daily Installments Report',
        rows=rows,
        summary=summary,
        start_date=start_date_str,
        end_date=end_date_str,
        loan_type_filter=loan_type_filter,
        loan_types=loan_types,
        today=today,
    )


@reports_bp.route('/export/daily-installments')
@login_required
@permission_required('view_reports')
def export_daily_installments():
    """Export daily installments report to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import date
    from decimal import Decimal

    today = date.today()
    start_date_str   = request.args.get('start_date', today.strftime('%Y-%m-%d'))
    end_date_str     = request.args.get('end_date',   today.strftime('%Y-%m-%d'))
    loan_type_filter = request.args.get('loan_type', '')

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date   = datetime.strptime(end_date_str,   '%Y-%m-%d').date()
    except ValueError:
        start_date = end_date = today

    loan_query = Loan.query.filter_by(status='active')
    loan_branch_filter = get_branch_filter_for_query(Loan.branch_id)
    if loan_branch_filter is not None:
        loan_query = loan_query.filter(loan_branch_filter)
    if loan_type_filter:
        loan_query = loan_query.filter_by(loan_type=loan_type_filter)
    loans = loan_query.all()

    from app.models import Customer as CustomerModel
    all_customer_ids = set()
    for loan in loans:
        if loan.guarantor_ids:
            for gid in loan.guarantor_ids.split(','):
                gid = gid.strip()
                if gid:
                    all_customer_ids.add(int(gid))
    guarantor_map = {}
    if all_customer_ids:
        gs = CustomerModel.query.filter(CustomerModel.id.in_(all_customer_ids)).all()
        guarantor_map = {g.id: g for g in gs}

    rows = []
    for loan in loans:
        schedule = loan.generate_payment_schedule()
        payments = loan.payments.order_by(LoanPayment.payment_date.desc()).all()
        guarantors = []
        if loan.guarantor_ids:
            for gid in loan.guarantor_ids.split(','):
                gid = gid.strip()
                if gid and int(gid) in guarantor_map:
                    guarantors.append(guarantor_map[int(gid)])

        for inst in schedule:
            due = inst['due_date']
            if not (start_date <= due <= end_date):
                continue
            if inst.get('is_skipped'):
                continue
            rows.append({
                'loan': loan,
                'due_date': due,
                'installment_number': inst['installment_number'],
                'amount': inst['amount'],
                'principal': inst['principal'],
                'interest': inst['interest'],
                'paid_amount': inst['paid_amount'],
                'remaining_amount': inst['remaining_amount'],
                'status': inst['status'],
                'guarantors': guarantors,
                'payments': payments,
            })

    rows.sort(key=lambda r: (r['due_date'], r['loan'].loan_number))

    wb = openpyxl.Workbook()

    # ── Sheet 1: Installments ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Installments'

    hdr_fill  = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    alt_fill  = PatternFill(start_color='E9F0F8', end_color='E9F0F8', fill_type='solid')
    ovd_fill  = PatternFill(start_color='FFDAD6', end_color='FFDAD6', fill_type='solid')
    paid_fill = PatternFill(start_color='D6F4D6', end_color='D6F4D6', fill_type='solid')
    ctr = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers1 = [
        'Due Date', 'Loan #', 'Loan Type', 'Inst #',
        'Customer Name', 'Member ID', 'Phone', 'NIC', 'Address',
        'Loan Amount', 'Disbursed', 'First Install Date', 'Maturity Date',
        'Installment Amt', 'Principal', 'Interest',
        'Paid', 'Remaining', 'Status',
        'Guarantor 1', 'G1 Phone', 'G1 NIC',
        'Guarantor 2', 'G2 Phone', 'G2 NIC',
        'Referred By', 'Staff Approver', 'Manager Approver', 'Final Approver',
    ]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = ctr
        cell.border = border

    for i, r in enumerate(rows, start=2):
        loan = r['loan']
        gs   = r['guarantors']
        g1 = gs[0] if len(gs) > 0 else None
        g2 = gs[1] if len(gs) > 1 else None
        status = r['status']

        row_data = [
            r['due_date'].strftime('%Y-%m-%d'),
            loan.loan_number,
            loan.loan_type or '',
            r['installment_number'],
            loan.customer.full_name if loan.customer else '',
            loan.customer.customer_id if loan.customer else '',
            loan.customer.phone_primary if loan.customer else '',
            loan.customer.nic_number if loan.customer else '',
            f"{loan.customer.address_line1 or ''}, {loan.customer.city or ''}" if loan.customer else '',
            float(loan.loan_amount or 0),
            float(loan.disbursed_amount or 0),
            loan.first_installment_date.strftime('%Y-%m-%d') if loan.first_installment_date else '',
            loan.maturity_date.strftime('%Y-%m-%d') if loan.maturity_date else '',
            float(r['amount']),
            float(r['principal']),
            float(r['interest']),
            float(r['paid_amount']),
            float(r['remaining_amount']),
            status.upper(),
            g1.full_name if g1 else '',
            g1.phone_primary if g1 else '',
            g1.nic_number if g1 else '',
            g2.full_name if g2 else '',
            g2.phone_primary if g2 else '',
            g2.nic_number if g2 else '',
            loan.referrer.full_name if loan.referrer else '',
            loan.staff_approver.full_name if loan.staff_approver else '',
            loan.manager_approver.full_name if loan.manager_approver else '',
            loan.final_approver.full_name if loan.final_approver else '',
        ]
        ws1.append(row_data)
        row_fill = ovd_fill if status == 'overdue' else (paid_fill if status == 'paid' else (alt_fill if i % 2 == 0 else None))
        for cell in ws1[i]:
            cell.border = border
            if row_fill:
                cell.fill = row_fill

    for col in ws1.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    # ── Sheet 2: Payment History ──────────────────────────────────────────
    ws2 = wb.create_sheet('Payment History')
    hdr2 = ['Loan #', 'Customer', 'Payment Date', 'Receipt #', 'Amount', 'Principal', 'Interest', 'Penalty', 'Balance After', 'Method', 'Collected By']
    ws2.append(hdr2)
    for cell in ws2[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = ctr
        cell.border = border

    seen_loans = {r['loan'].id for r in rows}
    seen_row = 2
    for r in rows:
        if r['loan'].id not in seen_loans:
            continue
        seen_loans.discard(r['loan'].id)  # only write each loan's history once
        for p in r['payments']:
            ws2.append([
                r['loan'].loan_number,
                r['loan'].customer.full_name if r['loan'].customer else '',
                p.payment_date.strftime('%Y-%m-%d') if p.payment_date else '',
                p.receipt_number or '',
                float(p.payment_amount or 0),
                float(p.principal_amount or 0),
                float(p.interest_amount or 0),
                float(p.penalty_amount or 0),
                float(p.balance_after or 0),
                p.payment_method or '',
                p.collected_by_user.full_name if p.collected_by else '',
            ])
            for cell in ws2[seen_row]:
                cell.border = border
                if seen_row % 2 == 0:
                    cell.fill = alt_fill
            seen_row += 1

    for col in ws2.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 3, 35)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    fname = f'daily_installments_{start_date_str}_to_{end_date_str}.xlsx'
    response.headers['Content-Disposition'] = f'attachment; filename={fname}'
    return response
